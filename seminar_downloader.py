#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
국회 정책세미나 VOD 다운로더

[설명]
https://vplatform.assembly.go.kr/ 사이트의 정책세미나 VOD를
GUI 환경에서 목록을 보고 다운로드하는 프로그램입니다.

[요구사항]
1. Python 3.x
2. requests 라이브러리 (pip install requests)
3. ffmpeg 프로그램 (시스템 PATH에 등록되어 있어야 함)
   - Windows: https://ffmpeg.org/download.html 에서 다운로드 후 환경변수 설정
   - macOS: brew install ffmpeg
   - Linux: sudo apt-get install ffmpeg (또는 배포판에 맞는 패키지 사용)

[실행]
터미널에서 'python seminar_downloader.py' 명령으로 실행해주세요.
오류 발생 시 터미널의 모든 출력을 복사해서 알려주시면 문제 해결에 도움이 됩니다.
"""
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import os
import json
import threading
import requests
import re
import subprocess
import sys
import tempfile
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta
import logging
import time # Added for Selenium waits

# Import the new Selenium extractor
from selenium_subtitle_extractor import SeleniumSubtitleExtractor # Added


# --- Constants ---
APP_VERSION = "2.4.4"
GITHUB_REPO = "doogiekang/Dodio_downloader"
UPDATE_CHECK_URL = f"https://api.github.com/repos/{GITHUB_REPO}/releases/latest"
UPDATE_INSTALLER_URL = ""  # 최신 릴리즈에서 자동으로 가져옴

DOWNLOAD_DIR    = os.path.join(os.path.expanduser("~"), "seminars_download")
JSON_FILE       = os.path.join(DOWNLOAD_DIR, "all_seminars.json")
SUMMARY_JSON_FILE = os.path.join(DOWNLOAD_DIR, "all_summaries.json")
W3_COMM_JSON    = os.path.join(DOWNLOAD_DIR, "w3_committees.json")
W3_CONF_JSON    = os.path.join(DOWNLOAD_DIR, "w3_conferences.json")
CONFIG_FILE     = os.path.join(DOWNLOAD_DIR, "config.json")
os.makedirs(DOWNLOAD_DIR, exist_ok=True)

# --- W3 영상회의록시스템 (상임위) ---
W3_BASE = "https://w3.assembly.go.kr"
W3_HEADERS = {
    'Accept': 'application/json, text/plain, */*',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
                  '(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Referer': 'https://w3.assembly.go.kr/main/index.do',
}

# 위원회명 키워드 → STT 서버 디렉토리명 매핑 (숫자 suffix 제외한 stem 기준)
_COMM_DIR_KEYWORDS = [
    ('gyoyuk',   ['교육']),
    ('gwabang',  ['과학기술정보방송통신', '과방', '과학기술', '방송통신']),
    ('beopsa',   ['법제사법', '법사']),
    ('gukto',    ['국토교통', '국토']),
    ('hwanno',   ['기후에너지환경노동', '환경노동', '환노', '기후에너지']),
    ('jeongmu',  ['정무']),
    ('munche',   ['문화체육관광', '문체']),
    ('bokji',    ['보건복지', '복지']),
    ('jaejeong', ['기획재정', '재정경제기획', '재정']),
    ('sangi',    ['산업통상', '중소벤처', '산자']),
    ('gukbang',  ['국방']),
    ('oetong',   ['외교통일', '외통']),
    ('haengan',  ['행정안전', '행안']),
    ('nonghae',  ['농림축산식품해양수산', '농해수', '농림']),
    ('unyeong',  ['국회운영', '운영위']),
    ('yegyeol',  ['예산결산', '예결']),
    ('woman',    ['여성가족', '여가', '성평등']),
    ('special',  ['특별위']),
    ('bon',      ['본회의']),
]

# --- Setup Logging ---
LOG_FILE = os.path.join(DOWNLOAD_DIR, "seminar_downloader.log")
_log_handler = logging.FileHandler(LOG_FILE, mode='w', encoding='utf-8')
_log_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
logging.getLogger().setLevel(logging.DEBUG)
logging.getLogger().addHandler(_log_handler)

class SeminarGUI:
    def __init__(self, master):
        logging.debug("SeminarGUI __init__ started.")
        self.master = master
        master.title("두디오 다운로더")
        master.geometry("1280x860")
        self._setup_styles()

        # --- 메뉴바 ---
        menubar = tk.Menu(master)
        settings_menu = tk.Menu(menubar, tearoff=0)
        settings_menu.add_command(label="Selenium 설정...", command=self._open_settings_dialog)
        settings_menu.add_command(label="STT 서버 설정...", command=self._open_stt_settings_dialog)
        menubar.add_cascade(label="설정", menu=settings_menu)
        master.config(menu=menubar)


        self.seminars = [] # This will be the combined list for seminars_by_id
        self.seminars_full = [] # Stores 'FULL' content (seminars)
        self.seminars_press = [] # Stores 'PRESS' content (press conferences)
        self.seminars_summary = [] # Stores 'SUMMARY' content (summary videos)
        self.seminars_by_id = {}
        self.fetch_thread = None
        self.download_thread = None
        self.full_bulk_subtitle_thread = None
        self.summary_bulk_subtitle_thread = None
        self.press_bulk_subtitle_thread = None
        self.excel_download_thread = None
        self._excel_items = []  # 엑셀에서 파싱한 아이템 목록
        self.ffmpeg_ready = self._check_ffmpeg()
        
        # --- Selenium Configuration Variables ---
        self.selenium_driver_path = tk.StringVar(value="") # User can set this
        self.selenium_headless_mode = tk.BooleanVar(value=True) # Headless by default
        self.selenium_user_agent = tk.StringVar(value="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/100.0.4896.127 Safari/537.36")
        self.selenium_extractor = None # Will be initialized when needed

        # --- STT 서버 설정 ---
        self.stt_host      = tk.StringVar(value="")
        self.stt_user      = tk.StringVar(value="")
        self.stt_password  = tk.StringVar(value="")
        self.stt_base_path = tk.StringVar(value="")
        # 보조 서버 (주 서버를 점프호스트로 터널링)
        self.stt_host2      = tk.StringVar(value="")
        self.stt_user2      = tk.StringVar(value="")
        self.stt_password2  = tk.StringVar(value="")
        self.stt_base_path2 = tk.StringVar(value="")
        self._load_config()

        # --- W3 상임위 탭 상태 ---
        self.w3_committee_map = {}   # mc -> name
        self._w3_name_to_mc = {}     # name -> mc
        self.w3_conf_cache = []
        self.w3_conf_data = {}       # tree iid -> conf dict
        self.w3_movie_data = []      # 현재 클립 목록
        self.w3_sami_is = '0'
        self.w3_current_conf = None
        self.w3_fetch_thread = None
        self.w3_download_thread = None

        # --- 장바구니 ---
        self.download_cart = []
        self.cart_download_thread = None
        self.full_current_item = None
        self.press_current_item = None
        self.summary_current_item = None

        self.content_list_url = "https://vplatform.assembly.go.kr/api/api/cms/content/contentList"

        # Store headers globally for reuse in download function
        self.headers = {
            'Accept': 'application/json, text/plain, */*',
            'Accept-Language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
            'Connection': 'keep-alive',
            'Content-Type': 'application/json;charset=UTF-8',
            'Origin': 'https://vplatform.assembly.go.kr',
            'Referer': 'https://vplatform.assembly.go.kr/',
            'Sec-Fetch-Dest': 'empty',
            'Sec-Fetch-Mode': 'cors',
            'Sec-Fetch-Site': 'same-origin',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/100.0.4896.127 Safari/537.36', # Generic Chrome/Windows UA
            'sec-ch-ua': '" Not A;Brand";v="99", "Chromium";v="100", "Google Chrome";v="100"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"'
        }
        
        # --- Main UI: Tabs ---
        self.notebook = ttk.Notebook(master)
        self.notebook.pack(pady=(8, 0), padx=10, fill="both", expand=True)

        self.downloader_tab = ttk.Frame(self.notebook, padding="12")
        self.summary_tab = ttk.Frame(self.notebook, padding="12")
        self.press_conference_tab = ttk.Frame(self.notebook, padding="12")

        self.committee_tab = ttk.Frame(self.notebook, padding="12")
        self.cart_tab     = ttk.Frame(self.notebook, padding="12")
        self.excel_tab    = ttk.Frame(self.notebook, padding="12")

        self.notebook.add(self.downloader_tab, text='세미나')
        self.notebook.add(self.summary_tab, text='세미나-요약영상')
        self.notebook.add(self.press_conference_tab, text='기자회견')
        self.notebook.add(self.committee_tab, text='상임위')
        self.notebook.add(self.cart_tab, text='장바구니 🛒')
        self.notebook.add(self.excel_tab, text='엑셀 업로드 📋')

        # --- Build Downloader Tab ---
        self._create_downloader_tab(self.downloader_tab)

        # --- Build Summary Tab ---
        self._create_summary_tab(self.summary_tab)

        # --- Build Press Conference Tab ---
        self._create_press_conference_tab(self.press_conference_tab)

        # --- Build Committee Tab ---
        self._create_committee_tab(self.committee_tab)

        # --- Build Cart Tab ---
        self._create_cart_tab(self.cart_tab)

        # --- Build Excel Tab ---
        self._create_excel_tab(self.excel_tab)

        # --- 상임위 탭: 캐시 파일에서 로드 (탭 전환 시) ---
        self._committee_loaded = False
        self.notebook.bind("<<NotebookTabChanged>>", self._on_tab_changed)

        # 상임위 회의 목록 캐시 {mc: [conf_list]}
        self._w3_conf_json_cache = {}

        # --- Status Bar ---
        ttk.Separator(master, orient='horizontal').pack(side=tk.BOTTOM, fill=tk.X)
        self.status_bar = ttk.Label(master, text="준비. 로컬 목록을 불러옵니다.",
                                    style='Status.TLabel', anchor=tk.W)
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)
        
        self.load_local_seminars()
        if not self.ffmpeg_ready:
            self.update_status("경고: ffmpeg를 찾을 수 없습니다. 다운로드 기능이 비활성화됩니다.")
            messagebox.showwarning(
                "ffmpeg 없음",
                f"ffmpeg를 찾을 수 없어 다운로드 기능이 비활성화됩니다.\n\n"
                f"자세한 내용은 로그 파일을 확인하세요:\n{LOG_FILE}"
            )

        # 백그라운드에서 업데이트 확인 (시작 지연 없이)
        threading.Thread(target=self._check_for_updates, daemon=True).start()

        logging.debug("SeminarGUI __init__ finished.")

    def _setup_styles(self):
        import tkinter.font as tkfont

        # 플랫폼별 폰트
        if sys.platform == 'win32':
            FAMILY = 'Segoe UI'
        elif sys.platform == 'darwin':
            FAMILY = 'Helvetica Neue'
        else:
            FAMILY = 'DejaVu Sans'

        FONT       = (FAMILY, 13)
        FONT_BOLD  = (FAMILY, 13, 'bold')
        FONT_SMALL = (FAMILY, 11)

        # tkinter 기본 named font 전체 교체 (ttkbootstrap 내부 참조 포함)
        for fname in ('TkDefaultFont', 'TkTextFont', 'TkFixedFont',
                      'TkMenuFont', 'TkHeadingFont', 'TkCaptionFont',
                      'TkSmallCaptionFont', 'TkIconFont', 'TkTooltipFont'):
            try:
                f = tkfont.nametofont(fname)
                f.configure(family=FAMILY, size=13)
            except Exception:
                pass

        style = ttk.Style()

        # 버튼: 점선 포커스 테두리 제거 (Windows), 여백 확대
        style.configure('TButton',
                        font=FONT_BOLD, padding=(14, 7),
                        focusthickness=0, focuscolor='')
        style.map('TButton', focuscolor=[('focus', ''), ('!focus', '')])

        # 탭: 굵은 폰트, 여유있는 패딩
        style.configure('TNotebook.Tab', font=FONT_BOLD, padding=[18, 8])

        # Treeview: 줄 간격 넉넉하게
        style.configure('Treeview',         font=FONT, rowheight=36)
        style.configure('Treeview.Heading', font=FONT_BOLD, padding=(8, 8))

        # LabelFrame 제목
        style.configure('TLabelframe.Label', font=FONT_BOLD)

        # Entry / Combobox
        style.configure('TEntry',    font=FONT)
        style.configure('TCombobox', font=FONT)

        # 커스텀 라벨 스타일
        style.configure('Muted.TLabel',  font=FONT_SMALL)
        style.configure('Status.TLabel', font=FONT_SMALL, padding=(12, 5))

        self._ui_font      = FONT
        self._ui_font_bold = FONT_BOLD

    def _create_content_list_tab(self, parent_frame, content_type_filter):
        """세미나/기자회견 목록 탭의 UI를 생성합니다."""
        # ── Toolbar ──
        toolbar = ttk.Frame(parent_frame)
        toolbar.pack(fill=tk.X, pady=(0, 6))

        refresh_button = ttk.Button(toolbar, text="목록 새로고침",
                                    command=self.start_fetch_thread)
        refresh_button.pack(side=tk.LEFT, padx=(0, 6))

        open_folder_button = ttk.Button(toolbar, text="폴더 열기",
                                        command=self.open_download_folder)
        open_folder_button.pack(side=tk.LEFT, padx=(0, 6))

        progress_label = ttk.Label(toolbar, text="", style='Muted.TLabel')
        progress_label.pack(side=tk.LEFT, padx=(0, 6))

        progress_bar = ttk.Progressbar(toolbar, orient="horizontal",
                                       length=180, mode="determinate")
        progress_bar.pack(side=tk.LEFT)

        ttk.Separator(parent_frame, orient='horizontal').pack(fill=tk.X, pady=(0, 8))

        # ── Search bar ──
        search_frame = ttk.Frame(parent_frame)
        search_frame.pack(fill=tk.X, pady=(0, 8))

        ttk.Label(search_frame, text="검색").pack(side=tk.LEFT, padx=(0, 6))
        search_var = tk.StringVar()
        search_entry = ttk.Entry(search_frame, textvariable=search_var)
        search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 6))
        ttk.Button(search_frame, text="지우기",
                   command=lambda: search_var.set("")).pack(side=tk.LEFT)

        # ── Content list ──
        list_frame = ttk.Frame(parent_frame)
        list_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 8))

        tree = self.create_treeview(list_frame)

        # ── Download panel ──
        dl_panel = ttk.LabelFrame(parent_frame, text="다운로드", padding=(10, 6))
        dl_panel.pack(fill=tk.X)

        selected_title_label = ttk.Label(dl_panel, text="선택된 항목: 없음",
                                         wraplength=1100, anchor=tk.W,
                                         style='Muted.TLabel')
        selected_title_label.pack(fill=tk.X, pady=(0, 6))

        ctrl_row = ttk.Frame(dl_panel)
        ctrl_row.pack(fill=tk.X)

        ttk.Label(ctrl_row, text="화질 선택").pack(side=tk.LEFT, padx=(0, 4))
        quality_var = tk.StringVar(value='고화질')
        ttk.Combobox(ctrl_row, textvariable=quality_var,
                     values=['고화질', '중화질', '저화질'],
                     state='readonly', width=10).pack(side=tk.LEFT, padx=(0, 8))

        add_cart_btn = ttk.Button(ctrl_row, text="장바구니 추가",
                                  command=lambda ctf=content_type_filter: self._add_content_to_cart(ctf),
                                  state=tk.DISABLED)
        add_cart_btn.pack(side=tk.LEFT, padx=(0, 8))

        if content_type_filter == "FULL":
            subtitle_btn = ttk.Button(ctrl_row, text="자막 다운로드 (TXT)",
                                      command=lambda: self.start_download_subtitle_selenium_thread(self.full_current_item) if self.full_current_item else None,
                                      state=tk.DISABLED)
        elif content_type_filter == "SUMMARY":
            subtitle_btn = ttk.Button(ctrl_row, text="자막 다운로드 (TXT)",
                                      command=lambda: self.start_download_subtitle_selenium_thread(self.summary_current_item) if self.summary_current_item else None,
                                      state=tk.DISABLED)
        else:
            subtitle_btn = ttk.Button(ctrl_row, text="자막 다운로드 (TXT)",
                                      command=lambda: self.start_download_ai_subtitle_thread(self.press_current_item) if self.press_current_item else None,
                                      state=tk.DISABLED)
        subtitle_btn.pack(side=tk.LEFT)

        if content_type_filter == "FULL":
            self.downloader_tree = tree
            self.downloader_progress_label = progress_label
            self.downloader_progress_bar = progress_bar
            self.downloader_refresh_button = refresh_button
            self.downloader_selected_title_label = selected_title_label
            self.downloader_search_var = search_var
            self.full_quality_var = quality_var
            self.full_add_cart_btn = add_cart_btn
            self.full_subtitle_btn = subtitle_btn
        elif content_type_filter == "SUMMARY":
            self.summary_tree = tree
            self.summary_progress_label = progress_label
            self.summary_progress_bar = progress_bar
            self.summary_refresh_button = refresh_button
            self.summary_selected_title_label = selected_title_label
            self.summary_search_var = search_var
            self.summary_quality_var = quality_var
            self.summary_add_cart_btn = add_cart_btn
            self.summary_subtitle_btn = subtitle_btn
        elif content_type_filter == "PRESS":
            self.press_conference_tree = tree
            self.press_conference_progress_label = progress_label
            self.press_conference_progress_bar = progress_bar
            self.press_conference_refresh_button = refresh_button
            self.press_conference_selected_title_label = selected_title_label
            self.press_conference_search_var = search_var
            self.press_quality_var = quality_var
            self.press_add_cart_btn = add_cart_btn
            self.press_subtitle_btn = subtitle_btn

        search_var.trace_add('write', lambda *args, ctf=content_type_filter: self._apply_search_filter(ctf))
        tree.bind("<<TreeviewSelect>>", lambda event, ctf=content_type_filter: self.on_content_select(event, ctf))
    
    def _create_downloader_tab(self, parent):
        self._create_content_list_tab(parent, "FULL")
        bulk_frame = ttk.Frame(parent)
        bulk_frame.pack(fill=tk.X, pady=(6, 0))
        ttk.Separator(bulk_frame, orient='horizontal').pack(fill=tk.X, pady=(0, 6))
        self.full_bulk_btn = ttk.Button(
            bulk_frame, text="세미나 자막 일괄다운로드",
            command=self._start_full_bulk_subtitle_download)
        self.full_bulk_btn.pack(side=tk.LEFT)
        self.full_bulk_label = ttk.Label(bulk_frame, text="", style='Muted.TLabel')
        self.full_bulk_label.pack(side=tk.LEFT, padx=(10, 0))

    def _create_summary_tab(self, parent):
        self._create_content_list_tab(parent, "SUMMARY")
        # 일괄 자막 다운로드 버튼 추가
        bulk_frame = ttk.Frame(parent)
        bulk_frame.pack(fill=tk.X, pady=(6, 0))
        ttk.Separator(bulk_frame, orient='horizontal').pack(fill=tk.X, pady=(0, 6))
        self.summary_bulk_btn = ttk.Button(
            bulk_frame,
            text="요약영상 자막 일괄다운로드",
            command=self._start_summary_bulk_subtitle_download
        )
        self.summary_bulk_btn.pack(side=tk.LEFT)
        self.summary_bulk_label = ttk.Label(bulk_frame, text="", style='Muted.TLabel')
        self.summary_bulk_label.pack(side=tk.LEFT, padx=(10, 0))

    def _create_press_conference_tab(self, parent):
        self._create_content_list_tab(parent, "PRESS")
        bulk_frame = ttk.Frame(parent)
        bulk_frame.pack(fill=tk.X, pady=(6, 0))
        ttk.Separator(bulk_frame, orient='horizontal').pack(fill=tk.X, pady=(0, 6))
        self.press_bulk_btn = ttk.Button(
            bulk_frame, text="기자회견 자막 일괄다운로드",
            command=self._start_press_bulk_subtitle_download)
        self.press_bulk_btn.pack(side=tk.LEFT)
        self.press_bulk_label = ttk.Label(bulk_frame, text="", style='Muted.TLabel')
        self.press_bulk_label.pack(side=tk.LEFT, padx=(10, 0))

    def _create_committee_tab(self, parent):
        """상임위 탭 UI를 생성합니다 (w3.assembly.go.kr)."""
        # ── Toolbar ──
        toolbar = ttk.Frame(parent)
        toolbar.pack(fill=tk.X, pady=(0, 6))

        self.w3_refresh_btn = ttk.Button(toolbar, text="상임위 목록 새로고침",
                                         command=self._start_w3_committee_fetch)
        self.w3_refresh_btn.pack(side=tk.LEFT, padx=(0, 6))

        ttk.Button(toolbar, text="폴더 열기",
                   command=self.open_download_folder).pack(side=tk.LEFT, padx=(0, 6))

        self.w3_progress_bar = ttk.Progressbar(toolbar, orient="horizontal",
                                               length=180, mode="determinate")
        self.w3_progress_bar.pack(side=tk.LEFT)

        ttk.Separator(parent, orient='horizontal').pack(fill=tk.X, pady=(0, 8))

        # ── Filter row ──
        filt = ttk.Frame(parent)
        filt.pack(fill=tk.X, pady=(0, 8))

        ttk.Label(filt, text="상임위").pack(side=tk.LEFT, padx=(0, 4))
        self.w3_comm_var = tk.StringVar()
        self.w3_comm_combo = ttk.Combobox(filt, textvariable=self.w3_comm_var,
                                           state='readonly', width=26)
        self.w3_comm_combo.pack(side=tk.LEFT, padx=(0, 16))
        self.w3_comm_combo.bind('<<ComboboxSelected>>', self._on_w3_committee_selected)

        ttk.Label(filt, text="기간").pack(side=tk.LEFT, padx=(0, 4))
        self.w3_start_var = tk.StringVar(value="2024-01-01")
        ttk.Entry(filt, textvariable=self.w3_start_var, width=12).pack(side=tk.LEFT, padx=(0, 2))
        ttk.Label(filt, text="~").pack(side=tk.LEFT, padx=4)
        self.w3_end_var = tk.StringVar(value=datetime.today().strftime('%Y-%m-%d'))
        ttk.Entry(filt, textvariable=self.w3_end_var, width=12).pack(side=tk.LEFT, padx=(0, 10))

        self.w3_search_btn = ttk.Button(filt, text="검색",
                                        command=self._on_w3_search_clicked)
        self.w3_search_btn.pack(side=tk.LEFT)

        # ── Vertical split ──
        paned = ttk.PanedWindow(parent, orient=tk.VERTICAL)
        paned.pack(fill=tk.BOTH, expand=True)

        # Upper pane: conference list
        upper = ttk.Frame(paned)
        paned.add(upper, weight=3)

        cols = ("date", "title")
        self.w3_conf_tree = ttk.Treeview(upper, columns=cols, show="headings")
        self.w3_conf_tree.heading("date",  text="날짜",     anchor=tk.W)
        self.w3_conf_tree.heading("title", text="회의 제목", anchor=tk.W)
        self.w3_conf_tree.column("date",  width=160, stretch=False)
        self.w3_conf_tree.column("title", width=800)

        vsb = ttk.Scrollbar(upper, orient="vertical",
                             command=self.w3_conf_tree.yview)
        self.w3_conf_tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        self.w3_conf_tree.pack(fill=tk.BOTH, expand=True)
        self.w3_conf_tree.bind('<<TreeviewSelect>>', self._on_w3_conf_selected)

        # Lower pane: clip list + download buttons
        lower = ttk.Frame(paned)
        paned.add(lower, weight=1)

        clip_lf = ttk.LabelFrame(lower, text="클립 선택  (Ctrl+클릭으로 복수 선택)", padding=6)
        clip_lf.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))

        clip_vsb = ttk.Scrollbar(clip_lf, orient=tk.VERTICAL)
        self.w3_clip_listbox = tk.Listbox(clip_lf, selectmode=tk.EXTENDED,
                                           yscrollcommand=clip_vsb.set,
                                           font=(self._ui_font[0], 13),
                                           background='#ffffff',
                                           foreground='#2c3e50',
                                           selectbackground='#18bc9c',
                                           selectforeground='#ffffff',
                                           activestyle='none',
                                           relief='flat', borderwidth=0,
                                           height=5)
        clip_vsb.config(command=self.w3_clip_listbox.yview)
        clip_vsb.pack(side=tk.RIGHT, fill=tk.Y)
        self.w3_clip_listbox.pack(fill=tk.BOTH, expand=True)

        dl_frame = ttk.Frame(lower)
        dl_frame.pack(side=tk.LEFT, anchor=tk.N)

        ttk.Label(dl_frame, text="화질 선택").pack(fill=tk.X, pady=(0, 2))
        self.w3_quality_var = tk.StringVar(value='고화질')
        ttk.Combobox(dl_frame, textvariable=self.w3_quality_var,
                     values=['고화질', '중화질', '저화질'],
                     state='readonly', width=14).pack(fill=tk.X, pady=(0, 6))

        self.w3_dl_video_btn = ttk.Button(dl_frame, text="장바구니 추가",
                                           command=self._add_w3_to_cart,
                                           state=tk.DISABLED, width=16)
        self.w3_dl_video_btn.pack(fill=tk.X, pady=(0, 4))

        self.w3_dl_txt_btn = ttk.Button(dl_frame, text="자막 다운로드 (TXT)",
                                         command=lambda: self._start_w3_subtitle_download('txt'),
                                         state=tk.DISABLED, width=16)
        self.w3_dl_txt_btn.pack(fill=tk.X, pady=(0, 4))

        self.w3_dl_all_btn = ttk.Button(dl_frame, text="전체 자막 다운로드",
                                         command=self._start_w3_full_subtitle_download,
                                         state=tk.DISABLED, width=16)
        self.w3_dl_all_btn.pack(fill=tk.X)

    def create_treeview(self, parent):
        """Treeview 위젯 생성 및 설정"""
        cols = ("date", "title", "duration", "qualities")
        tree = ttk.Treeview(parent, columns=cols, show="headings", height=15)

        tree.heading("date", text="날짜", anchor=tk.W)
        tree.heading("title", text="세미나 제목", anchor=tk.W)
        tree.heading("duration", text="재생 시간", anchor=tk.W)
        tree.heading("qualities", text="제공화질", anchor=tk.W)

        tree.column("date", width=160, stretch=False)
        tree.column("title", width=700)
        tree.column("duration", width=100, stretch=False, anchor=tk.CENTER)
        tree.column("qualities", width=150, stretch=False, anchor=tk.CENTER)
        
        vsb = ttk.Scrollbar(parent, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(parent, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        hsb.pack(side=tk.BOTTOM, fill=tk.X)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        return tree

    def _populate_tree_for_tabs(self):
        logging.debug("_populate_tree_for_tabs started.")

        # 최신순 정렬
        self.seminars_full.sort(key=lambda x: x.get('date', ''), reverse=True)
        self.seminars_press.sort(key=lambda x: x.get('date', ''), reverse=True)
        self.seminars_summary.sort(key=lambda x: x.get('date', ''), reverse=True)

        # Clear existing items from all trees
        if hasattr(self, 'downloader_tree'):
            for i in self.downloader_tree.get_children():
                self.downloader_tree.delete(i)
        if hasattr(self, 'press_conference_tree'):
            for i in self.press_conference_tree.get_children():
                self.press_conference_tree.delete(i)
        if hasattr(self, 'summary_tree'):
            for i in self.summary_tree.get_children():
                self.summary_tree.delete(i)

        # Populate Downloader Tab
        downloader_inserted_count = 0
        for seminar in self.seminars_full:
            qualities = [v.get('vodRes', '?') for v in seminar.get('vod_list', [])]
            duration_sec = seminar.get('duration', 0)
            duration_str = self._format_duration(duration_sec)
            values = (seminar.get('date', 'N/A'), seminar.get('title', 'N/A'), duration_str, ' | '.join(qualities))
            self.downloader_tree.insert("", "end", values=values, tags=(seminar.get('id'),))
            downloader_inserted_count += 1
        logging.debug(f"Populated downloader_tree with {downloader_inserted_count} items.")

        # Populate Summary Tab
        summary_inserted_count = 0
        for item in self.seminars_summary:
            qualities = [v.get('vodRes', '?') for v in item.get('vod_list', [])]
            duration_sec = item.get('duration', 0)
            duration_str = self._format_duration(duration_sec)
            values = (item.get('date', 'N/A'), item.get('title', 'N/A'), duration_str, ' | '.join(qualities))
            self.summary_tree.insert("", "end", values=values, tags=(item.get('id'),))
            summary_inserted_count += 1
        logging.debug(f"Populated summary_tree with {summary_inserted_count} items.")

        # Populate Press Conference Tab
        press_inserted_count = 0
        for pc in self.seminars_press:
            qualities = [v.get('vodRes', '?') for v in pc.get('vod_list', [])]
            duration_sec = pc.get('duration', 0)
            duration_str = self._format_duration(duration_sec)
            values = (pc.get('date', 'N/A'), pc.get('title', 'N/A'), duration_str, ' | '.join(qualities))
            self.press_conference_tree.insert("", "end", values=values, tags=(pc.get('id'),))
            press_inserted_count += 1
        logging.debug(f"Populated press_conference_tree with {press_inserted_count} items.")

        self.update_status(f"총 세미나 {downloader_inserted_count}개, 요약영상 {summary_inserted_count}개, 기자회견 {press_inserted_count}개를 불러왔습니다.")
        logging.debug("_populate_tree_for_tabs finished.")

    def _apply_search_filter(self, content_type_filter):
        """검색어에 맞는 항목만 트리뷰에 표시합니다."""
        if content_type_filter == "FULL":
            tree = self.downloader_tree
            data_list = self.seminars_full
            query = self.downloader_search_var.get().strip().lower()
        elif content_type_filter == "SUMMARY":
            tree = self.summary_tree
            data_list = self.seminars_summary
            query = self.summary_search_var.get().strip().lower()
        elif content_type_filter == "PRESS":
            tree = self.press_conference_tree
            data_list = self.seminars_press
            query = self.press_conference_search_var.get().strip().lower()
        else:
            return

        for i in tree.get_children():
            tree.delete(i)

        for item in data_list:
            title = item.get('title', '')
            if query and query not in title.lower():
                continue
            qualities = [v.get('vodRes', '?') for v in item.get('vod_list', [])]
            duration_str = self._format_duration(item.get('duration', 0))
            values = (item.get('date', 'N/A'), title, duration_str, ' | '.join(qualities))
            tree.insert("", "end", values=values, tags=(item.get('id'),))

    def _format_duration(self, duration_sec):
        if duration_sec:
            mins, secs = divmod(duration_sec, 60)
            hours, mins = divmod(mins, 60)
            return f"{int(hours):02}:{int(mins):02}:{int(secs):02}" if hours > 0 else f"{int(mins):02}:{int(secs):02}"
        else:
            return "N/A"

    def on_content_select(self, event=None, content_type_filter=None):
        if content_type_filter == "FULL":
            current_tree = self.downloader_tree
            current_selected_title_label = self.downloader_selected_title_label
            add_cart_btn = self.full_add_cart_btn
            subtitle_btn = self.full_subtitle_btn
        elif content_type_filter == "SUMMARY":
            current_tree = self.summary_tree
            current_selected_title_label = self.summary_selected_title_label
            add_cart_btn = self.summary_add_cart_btn
            subtitle_btn = self.summary_subtitle_btn
        elif content_type_filter in ("PRESS", "PRESSCONF"):
            content_type_filter = "PRESS"
            current_tree = self.press_conference_tree
            current_selected_title_label = self.press_conference_selected_title_label
            add_cart_btn = self.press_add_cart_btn
            subtitle_btn = self.press_subtitle_btn
        else:
            return

        selected_items = current_tree.selection()
        if not selected_items:
            current_selected_title_label.config(text="선택: 없음")
            if content_type_filter == "FULL":
                self.full_current_item = None
            elif content_type_filter == "SUMMARY":
                self.summary_current_item = None
            else:
                self.press_current_item = None
            add_cart_btn.config(state=tk.DISABLED)
            subtitle_btn.config(state=tk.DISABLED)
            return

        item_id_str = current_tree.item(selected_items[0], 'tags')[0]
        content_item = self.seminars_by_id.get(item_id_str)

        if not content_item:
            current_selected_title_label.config(text="선택: 없음 (데이터 오류)")
            return

        if content_type_filter == "FULL":
            self.full_current_item = content_item
        elif content_type_filter == "SUMMARY":
            self.summary_current_item = content_item
        else:
            self.press_current_item = content_item

        current_selected_title_label.config(text=f"선택: {content_item.get('title', 'N/A')}")

        is_cart_downloading = self.cart_download_thread and self.cart_download_thread.is_alive()
        btn_state = tk.DISABLED if (not self.ffmpeg_ready or is_cart_downloading) else tk.NORMAL
        add_cart_btn.config(state=btn_state)
        subtitle_btn.config(state=tk.DISABLED if is_cart_downloading else tk.NORMAL)
        
    def _get_default_config_path(self):
        if getattr(sys, 'frozen', False):
            base = sys._MEIPASS
        else:
            base = os.path.dirname(os.path.abspath(__file__))
        return os.path.join(base, 'default_config.json')

    def _load_config(self):
        # 사용자 설정 우선, 없으면 번들된 기본값 사용
        if os.path.exists(CONFIG_FILE):
            cfg_path = CONFIG_FILE
        else:
            cfg_path = self._get_default_config_path()
            if not os.path.exists(cfg_path):
                return
        try:
            with open(cfg_path, 'r', encoding='utf-8') as f:
                cfg = json.load(f)
            self.stt_host.set(cfg.get('stt_host', ''))
            self.stt_user.set(cfg.get('stt_user', ''))
            self.stt_password.set(cfg.get('stt_password', ''))
            self.stt_base_path.set(cfg.get('stt_base_path', ''))
            self.stt_host2.set(cfg.get('stt_host2', ''))
            self.stt_user2.set(cfg.get('stt_user2', ''))
            self.stt_password2.set(cfg.get('stt_password2', ''))
            self.stt_base_path2.set(cfg.get('stt_base_path2', ''))
            logging.debug(f"Config loaded from: {cfg_path}")
        except Exception as e:
            logging.warning(f"Config load failed: {e}")

    def _save_config(self):
        try:
            cfg = {
                'stt_host':       self.stt_host.get(),
                'stt_user':       self.stt_user.get(),
                'stt_password':   self.stt_password.get(),
                'stt_base_path':  self.stt_base_path.get(),
                'stt_host2':      self.stt_host2.get(),
                'stt_user2':      self.stt_user2.get(),
                'stt_password2':  self.stt_password2.get(),
                'stt_base_path2': self.stt_base_path2.get(),
            }
            with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
                json.dump(cfg, f, ensure_ascii=False, indent=2)
            logging.debug("Config saved to file.")
        except Exception as e:
            logging.warning(f"Config save failed: {e}")

    def load_local_seminars(self):
        logging.debug("load_local_seminars started.")
        if os.path.exists(JSON_FILE):
            try:
                with open(JSON_FILE, 'r', encoding='utf-8') as f:
                    combined_seminars = json.load(f)
                    for item in combined_seminars:
                        content_type = item.get('content_type', 'FULL')
                        if content_type == "FULL":
                            self.seminars_full.append(item)
                        elif content_type in ("PRESS", "PRESSCONF"):
                            self.seminars_press.append(item)
                        self.seminars_by_id[item['id']] = item
                logging.debug(f"Loaded {len(self.seminars_full)} seminars and {len(self.seminars_press)} press conferences from local file.")
            except (json.JSONDecodeError, IOError) as e:
                self.update_status(f"로컬 파일 로드 오류: {e}")
                logging.error(f"Local file load error: {e}")
                messagebox.showerror("오류", f"저장된 세미나 파일({JSON_FILE})을 읽는 데 실패했습니다.")

        if os.path.exists(SUMMARY_JSON_FILE):
            try:
                with open(SUMMARY_JSON_FILE, 'r', encoding='utf-8') as f:
                    summary_list = json.load(f)
                    for item in summary_list:
                        self.seminars_summary.append(item)
                        self.seminars_by_id[item['id']] = item
                logging.debug(f"Loaded {len(self.seminars_summary)} summary videos from local file.")
            except (json.JSONDecodeError, IOError) as e:
                logging.error(f"Summary file load error: {e}")

        if not self.seminars_full and not self.seminars_press and not self.seminars_summary:
            self.update_status("저장된 목록이 없습니다. '새로고침'을 눌러주세요.")
        self._populate_tree_for_tabs()
        logging.debug("load_local_seminars finished.")

    def start_fetch_thread(self):
        logging.debug("start_fetch_thread called.")
        if self.fetch_thread and self.fetch_thread.is_alive():
            messagebox.showwarning("진행 중", "이미 세미나 목록을 가져오는 중입니다.")
            return
        
        # Use the correct refresh button based on the active tab, though for initial fetch, it's global
        self.downloader_refresh_button.config(state=tk.DISABLED) # Changed refresh_button to downloader_refresh_button
        self.fetch_thread = threading.Thread(target=self._fetch_seminar_data, daemon=True)
        self.fetch_thread.start()
        logging.debug("Fetch thread started.")

    def _fetch_seminar_data(self):
        logging.debug("_fetch_seminar_data started (parallel fetch).")
        self.seminars_full = []
        self.seminars_press = []
        self.seminars_summary = []
        self.seminars_by_id = {}

        results = {}

        def fetch_one(content_type):
            results[content_type] = self._fetch_one_content_type(content_type)

        threads = [
            threading.Thread(target=fetch_one, args=('FULL',), daemon=True),
            threading.Thread(target=fetch_one, args=('PRESSCONF',), daemon=True),
            threading.Thread(target=fetch_one, args=('SUMMARY',), daemon=True),
        ]
        for t in threads:
            t.start()
        for t in threads:
            t.join()

        self.seminars_full = results.get('FULL', [])
        self.seminars_press = results.get('PRESSCONF', [])
        self.seminars_summary = results.get('SUMMARY', [])
        for item in self.seminars_full + self.seminars_press + self.seminars_summary:
            self.seminars_by_id[item['id']] = item

        total = len(self.seminars_full) + len(self.seminars_press) + len(self.seminars_summary)
        try:
            combined_seminars = self.seminars_full + self.seminars_press
            with open(JSON_FILE, 'w', encoding='utf-8') as f:
                json.dump(combined_seminars, f, ensure_ascii=False, indent=2)
            with open(SUMMARY_JSON_FILE, 'w', encoding='utf-8') as f:
                json.dump(self.seminars_summary, f, ensure_ascii=False, indent=2)
            self.update_status(f"완료! 총 {total}개 항목을 찾아 저장했습니다.")
            logging.debug(f"Saved {total} items (summary: {len(self.seminars_summary)})")
        except IOError as e:
            self.update_status(f"파일 저장 오류: {e}")
            logging.error(f"File save error: {e}", exc_info=True)
            self.master.after(0, messagebox.showerror, "오류", f"결과를 파일에 저장하는데 실패했습니다.")

        self.master.after(0, self._populate_tree_for_tabs)
        self.master.after(0, self.downloader_refresh_button.config, {'state': tk.NORMAL})
        logging.debug("_fetch_seminar_data finished.")

    def _fetch_one_content_type(self, content_type):
        """날짜 범위 적응형 방식으로 전체 목록을 가져옵니다.
        서버가 page 파라미터를 무시하므로 startDt/endDt 필터로 분할 수집합니다."""
        label = {"FULL": "세미나", "SUMMARY": "요약영상", "PRESSCONF": "기자회견"}.get(content_type, content_type)
        all_items = []
        unique_ids = set()
        today = datetime.today()

        for year in range(2000, today.year + 1):
            year_start = datetime(year, 1, 1)
            year_end = min(datetime(year, 12, 31), today)
            self._fetch_date_range_recursive(
                content_type, label, all_items, unique_ids, year_start, year_end
            )

        all_items.sort(key=lambda x: x.get('date', ''), reverse=True)
        logging.debug(f"{content_type} fetch complete: {len(all_items)} items")
        return all_items

    def _fetch_date_range_recursive(self, content_type, label, all_items, unique_ids, start_dt, end_dt):
        """날짜 범위로 데이터를 가져오고, 500건이면 절반으로 나눠 재귀 호출합니다."""
        MAX_PAGE_SIZE = 500
        start_str = start_dt.strftime('%Y-%m-%d')
        end_str = end_dt.strftime('%Y-%m-%d')

        payload = json.dumps({
            "fcltNameList": [], "policyIdList": [],
            "contentType": content_type,
            "page": 0, "pageSize": MAX_PAGE_SIZE,
            "startDt": start_str, "endDt": end_str,
            "monaCdList": []
        })

        items = []
        for attempt in range(3):
            try:
                response = requests.post(self.content_list_url, headers=self.headers, data=payload, timeout=30)
                response.raise_for_status()
                items = response.json().get('data', [])
                break
            except requests.exceptions.ConnectionError:
                if attempt < 2:
                    time.sleep(1.5)
                else:
                    return
            except Exception as e:
                logging.error(f"Error fetching {content_type} {start_str}~{end_str}: {e}")
                return

        if len(items) < MAX_PAGE_SIZE or (end_dt - start_dt).days <= 1:
            # 500건 미만이거나 더 이상 분할 불가 → 결과 추가
            for item in items:
                cid = item.get('cid')
                if not cid or str(cid) in unique_ids:
                    continue

                duration_raw = item.get('duration', '')
                duration_in_seconds = 0
                if duration_raw and str(duration_raw).startswith('T'):
                    parts = str(duration_raw)[1:].split(':')
                    try:
                        if len(parts) == 3:
                            h, m, s = map(int, parts)
                            duration_in_seconds = h * 3600 + m * 60 + s
                        elif len(parts) == 2:
                            m, s = map(int, parts)
                            duration_in_seconds = m * 60 + s
                    except (ValueError, IndexError):
                        pass

                vod_list = []
                if item.get('videoFile1'): vod_list.append({'vodUrl': item['videoFile1'], 'vodRes': 'High'})
                if item.get('videoFile2'): vod_list.append({'vodUrl': item['videoFile2'], 'vodRes': 'Medium'})
                if item.get('videoFile3'): vod_list.append({'vodUrl': item['videoFile3'], 'vodRes': 'Low'})

                all_items.append({
                    'id': str(cid),
                    'sid': str(item.get('sid', '')),
                    'title': item.get('seminarTitle', item.get('title', 'N/A')).replace('&#8230;', '...').replace('&#34;', '"').replace('&#39;', "'").replace('&#183;', '·'),
                    'date': item.get('eventDt', '')[:10] + ' ' + item.get('beginTime', ''),
                    'duration': duration_in_seconds,
                    'vod_list': vod_list,
                    'subtitle_url': item.get('smiUrl') or item.get('vttUrl'),
                    'content_type': content_type,
                })
                unique_ids.add(str(cid))

            self.update_status(f"{label} {len(all_items)}건 수집 중 ({start_str} ~ {end_str})")
            time.sleep(0.05)
        else:
            # 500건 → 절반으로 분할해 재귀 호출
            mid = start_dt + (end_dt - start_dt) // 2
            self._fetch_date_range_recursive(content_type, label, all_items, unique_ids, start_dt, mid)
            self._fetch_date_range_recursive(content_type, label, all_items, unique_ids, mid + timedelta(days=1), end_dt)

    def start_download_thread(self, url, seminar_data):
        logging.debug(f"start_download_thread called for url: {url}")
        if self.download_thread and self.download_thread.is_alive():
            messagebox.showwarning("진행 중", "이미 다른 파일을 다운로드 중입니다.")
            return

        if not seminar_data:
            messagebox.showerror("오류", "다운로드 정보가 올바르지 않습니다.")
            return
        # PRESSCONF는 contentInfo API로 URL을 받아오므로 file_key(videoFile1/2/3) 허용
        if not url and seminar_data.get('content_type') != 'PRESSCONF':
            messagebox.showerror("오류", "다운로드 URL이 없습니다.")
            return

        self.download_thread = threading.Thread(target=self._download_video, args=(url, seminar_data), daemon=True)
        self.download_thread.start()
        logging.debug("Download thread started.")

    def _download_video(self, url, seminar, subtitle_only=False):
        logging.debug(f"_download_video started for seminar ID: {seminar.get('id')}, url: {url}")
        ct = seminar.get('content_type', 'FULL')
        if ct == 'FULL':
            progress_bar = self.downloader_progress_bar
        elif ct == 'SUMMARY':
            progress_bar = self.summary_progress_bar
        else:
            progress_bar = self.press_conference_progress_bar
        self.master.after(0, lambda ct=seminar['content_type']: self.on_content_select(event=None, content_type_filter=ct))

        title = self._sanitize_filename(seminar['title'])
        base_filename = os.path.join(DOWNLOAD_DIR, f"{self._date_prefix(seminar)}{title}")
        video_filename = f"{base_filename}.mp4"
        subtitle_url = seminar.get('subtitle_url')
        ai_cc_id = None

        # contentInfo API로 선택 화질 URL 취득
        # PRESSCONF는 항상, FULL은 url 자리에 file_key('videoFile1/2/3')가 전달된 경우
        if seminar.get('content_type') == 'PRESSCONF' or url in ('videoFile1', 'videoFile2', 'videoFile3'):
            file_key = url if url in ('videoFile1', 'videoFile2', 'videoFile3') else 'videoFile1'
            quality_label = {'videoFile1': '고화질', 'videoFile2': '중화질', 'videoFile3': '저화질'}
            content_type = seminar.get('content_type', 'FULL')
            self.update_status(f"화질 정보 조회 중 ({quality_label.get(file_key)}): {title}...")
            info = self._fetch_content_info(seminar['id'], seminar.get('sid', '0'), content_type)
            if info:
                url = info.get(file_key) or info.get('videoFile1') or url
                ai_cc_id = info.get('aiCC')
                logging.debug(f"contentInfo: type={content_type}, file_key={file_key}, url={url}, aiCC={ai_cc_id}")
            else:
                self.update_status("화질 정보 조회 실패.")
                url = ''

        if not url:
            self.update_status("오류: 다운로드 URL이 없습니다.")
            self.master.after(0, lambda ct=seminar['content_type']: self.on_content_select(event=None, content_type_filter=ct))
            return

        # ── 자막만 모드: 영상 다운로드 건너뜀 ──
        if subtitle_only:
            self._download_subtitle_only(subtitle_url, ai_cc_id, base_filename, title)
            self.master.after(0, lambda ct=seminar['content_type']: self.on_content_select(event=None, content_type_filter=ct))
            return

        if os.path.exists(video_filename):
            self.update_status(f"이미 존재함: {os.path.basename(video_filename)}")
            self.master.after(0, lambda ct=seminar['content_type']: self.on_content_select(event=None, content_type_filter=ct))
            return

        # Construct the HLS playlist URL
        m3u8_url = f"{url}/playlist.m3u8"
        logging.debug(f"Constructed HLS playlist URL: {m3u8_url}")

        try:
            self.update_status(f"비디오 다운로드 중: {os.path.basename(video_filename)}...")
            self.master.after(0, progress_bar.config, {'mode': 'indeterminate'})
            self.master.after(0, progress_bar.start)

            # Use ffmpeg to directly download and mux the HLS stream
            cmd = [self._get_ffmpeg_path(),
                   '-reconnect', '1', '-reconnect_streamed', '1', '-reconnect_delay_max', '30',
                   '-i', m3u8_url,
                   '-c', 'copy', '-bsf:a', 'aac_adtstoasc', '-y', '-loglevel', 'error', video_filename]

            # creationflags는 Windows 전용 (콘솔 창 숨김)
            extra_kwargs = {}
            if sys.platform == "win32":
                extra_kwargs['creationflags'] = subprocess.CREATE_NO_WINDOW

            process = subprocess.run(cmd, check=True, capture_output=True, text=True, encoding='utf-8', **extra_kwargs)

            self.master.after(0, progress_bar.stop)
            self.master.after(0, progress_bar.config, {'mode': 'determinate', 'value': 0})

            if os.path.exists(video_filename):
                size = os.path.getsize(video_filename) / (1024 * 1024)
                self.update_status(f"비디오 다운로드 완료: {os.path.basename(video_filename)} ({size:.1f}MB)")
                logging.debug(f"Video download complete: {video_filename} ({size:.1f}MB)")

                # 자막 다운로드
                self._download_subtitle_only(subtitle_url, ai_cc_id, base_filename, title)

            else:
                 self.update_status(f"비디오 다운로드 실패: {os.path.basename(video_filename)}")
                 logging.error(f"Video download failed, file not found after ffmpeg: {video_filename}")
        except subprocess.CalledProcessError as e:
            error_msg = e.stderr
            user_error_message = "알 수 없는 ffmpeg 오류"
            if "Server returned 404 Not Found" in error_msg:
                user_error_message = "비디오 파일을 찾을 수 없습니다 (Server returned 404 Not Found). 링크가 만료되었거나 잘못된 것 같습니다."
            elif error_msg.strip():
                user_error_message = error_msg.strip().splitlines()[-1]
            
            self.master.after(0, progress_bar.stop)
            self.master.after(0, progress_bar.config, {'mode': 'determinate', 'value': 0})
            self.master.after(0, self.update_status, f"비디오 다운로드 오류: {user_error_message}")
            logging.error(f"ffmpeg CalledProcessError: {e.returncode}, Stderr: {error_msg}", exc_info=True)
            self.master.after(0, messagebox.showerror, "다운로드 오류", f"비디오 다운로드 중 오류가 발생했습니다:\n{user_error_message}")
        except Exception as e:
            self.master.after(0, progress_bar.stop)
            self.master.after(0, progress_bar.config, {'mode': 'determinate', 'value': 0})
            self.master.after(0, self.update_status, f"비디오 다운로드 중 예외 발생: {e}")
            logging.error(f"An unhandled exception occurred in _download_video: {e}", exc_info=True)
            messagebox.showerror("치명적 오류", f"프로그램 실행 중 예기치 않은 오류가 발생했습니다. 자세한 내용은 {LOG_FILE} 파일을 확인해주세요.\n\n오류: {e}")
        self.master.after(0, lambda ct=seminar['content_type']: self.on_content_select(event=None, content_type_filter=ct))
        logging.debug("_download_video finished.")

    def _download_subtitle_only(self, subtitle_url, ai_cc_id, base_filename, title):
        """자막 URL 또는 aiCC ID로 자막만 다운로드합니다."""
        if subtitle_url:
            ext = ".smi" if "smi" in subtitle_url.lower() else (".vtt" if "vtt" in subtitle_url.lower() else ".txt")
            subtitle_filename = f"{base_filename}{ext}"
            try:
                self.update_status(f"자막 다운로드 중: {os.path.basename(subtitle_filename)}...")
                r = requests.get(subtitle_url, headers=self.headers, timeout=10)
                r.raise_for_status()
                with open(subtitle_filename, 'wb') as f:
                    f.write(r.content)
                self.update_status(f"자막 완료: {os.path.basename(subtitle_filename)}")
            except Exception as e:
                self.update_status(f"자막 다운로드 실패: {e}")
                logging.error(f"Subtitle download failed {subtitle_url}: {e}", exc_info=True)
        elif ai_cc_id:
            try:
                self.update_status(f"AI 자막 다운로드 중: {title}...")
                result = self._download_ai_subtitle(ai_cc_id, base_filename)
                if result:
                    self.update_status(f"AI 자막 완료: {os.path.basename(result)}")
                else:
                    self.update_status("자막 없음")
            except Exception as e:
                logging.warning(f"AI 자막 다운로드 실패 (무시): {e}")
                self.update_status("자막 없음")
        else:
            self.update_status(f"자막 없음: {title}")

    # ──────────────────────────────────────────────────────────────
    # PRESSCONF 전용: contentInfo / AI 자막
    # ──────────────────────────────────────────────────────────────

    def _fetch_content_info(self, cid, sid, content_type):
        """contentInfo API를 호출해 전체 화질 URL과 aiCC를 반환합니다."""
        url = "https://vplatform.assembly.go.kr/api/api/cms/content/contentInfo"
        payload = json.dumps({
            "cid": int(cid),
            "sid": int(sid) if str(sid).isdigit() else 0,
            "contentType": content_type
        })
        try:
            response = requests.post(url, headers=self.headers, data=payload, timeout=30)
            response.raise_for_status()
            data = response.json()
            if data.get('code') == 200 and data.get('data'):
                result = data['data']
                # data가 list인 경우 첫 번째 항목 사용
                outer = result[0] if isinstance(result, list) else result
                # 실제 영상 정보는 outer["data"] 안에 중첩되어 있고
                # aiCC는 outer 레벨에 있음 → 병합해서 반환
                inner = outer.get('data') or {}
                if isinstance(inner, dict):
                    inner = dict(inner)  # 원본 변경 방지
                    inner['aiCC'] = outer.get('aiCC')
                    logging.debug(f"contentInfo inner keys for cid={cid}: {list(inner.keys())}")
                    return inner
                logging.warning(f"contentInfo inner data is not a dict for cid={cid}: {type(inner)}")
            logging.warning(f"contentInfo returned non-200 or empty data for cid={cid}: {data.get('code')}")
        except Exception as e:
            logging.error(f"contentInfo API error for cid={cid}: {e}", exc_info=True)
        return None

    def _download_ai_subtitle(self, ai_cc_id, base_filename):
        """aiCC ID로 videometaapi에서 자막을 받아 TXT 파일로 저장합니다."""
        subtitle_api_url = "https://videometaapi.assembly.go.kr/metaApi/api/subtitles/apiContentsSubtitleView.do"
        params = {"maxLength": 17, "contId": ai_cc_id}
        subtitle_headers = dict(self.headers)
        subtitle_headers["Referer"] = "https://openplayer.assembly.go.kr/"
        subtitle_headers["Origin"] = "https://openplayer.assembly.go.kr"

        response = requests.get(subtitle_api_url, headers=subtitle_headers, params=params, timeout=30)
        response.raise_for_status()
        data = response.json()

        smi_list = data.get('smiList', [])
        if not smi_list:
            logging.warning(f"smiList is empty for aiCC={ai_cc_id}")
            return None

        txt_content = self._smilist_to_txt(smi_list)
        txt_filename = f"{base_filename}.txt"
        with open(txt_filename, 'w', encoding='utf-8') as f:
            f.write(txt_content)
        logging.debug(f"AI subtitle saved as txt: {txt_filename} ({len(smi_list)} entries)")
        return txt_filename

    @staticmethod
    def _smilist_to_txt(smi_list):
        """smiList JSON의 cc 텍스트만 추출해 순수 텍스트로 반환합니다."""
        lines = []
        for item in smi_list:
            text = item.get('cc', '').strip()
            if text:
                lines.append(text)
        return '\n'.join(lines)

    @staticmethod
    def _smilist_to_srt(smi_list):
        """smiList JSON(cc/start/end)을 SRT 형식 문자열로 변환합니다."""
        def to_ts(seconds):
            h = int(seconds // 3600)
            m = int((seconds % 3600) // 60)
            s = int(seconds % 60)
            ms = int(round((seconds % 1) * 1000))
            return f"{h:02d}:{m:02d}:{s:02d},{ms:03d}"

        lines = []
        idx = 1
        for item in smi_list:
            text = item.get('cc', '').strip()
            if not text:
                continue
            start = to_ts(item.get('start', 0))
            end = to_ts(item.get('end', 0))
            lines.append(f"{idx}\n{start} --> {end}\n{text}\n")
            idx += 1
        return '\n'.join(lines)

    def start_download_ai_subtitle_thread(self, seminar_data):
        """자막만 다운로드하는 스레드를 시작합니다 (PRESSCONF 전용)."""
        logging.debug(f"start_download_ai_subtitle_thread called for cid={seminar_data.get('id')}")
        if self.download_thread and self.download_thread.is_alive():
            messagebox.showwarning("진행 중", "이미 다운로드 중입니다.")
            return
        if not seminar_data:
            messagebox.showerror("오류", "다운로드 정보가 올바르지 않습니다.")
            return
        self.download_thread = threading.Thread(
            target=self._download_ai_subtitle_only, args=(seminar_data,), daemon=True
        )
        self.download_thread.start()

    def _download_ai_subtitle_only(self, seminar_data):
        """PRESSCONF 자막만 다운로드합니다 (영상 다운로드 없음)."""
        cid = seminar_data.get('id')
        title = self._sanitize_filename(seminar_data['title'])
        base_filename = os.path.join(DOWNLOAD_DIR, f"{self._date_prefix(seminar_data)}{title}")
        progress_bar = self.press_conference_progress_bar

        self.master.after(0, lambda ct=seminar_data['content_type']: self.on_content_select(event=None, content_type_filter=ct))
        self.master.after(0, progress_bar.config, {'mode': 'indeterminate'})
        self.master.after(0, progress_bar.start)

        try:
            self.update_status(f"기자회견 자막 정보 조회 중: {title}...")
            info = self._fetch_content_info(cid, seminar_data.get('sid', '0'), 'PRESSCONF')
            if not info or not info.get('aiCC'):
                self.update_status("자막 정보를 찾을 수 없습니다 (aiCC 없음).")
                logging.warning(f"No aiCC found for cid={cid}")
                return

            ai_cc_id = info['aiCC']
            self.update_status(f"AI 자막 다운로드 중 (contId={ai_cc_id}): {title}...")
            srt_path = self._download_ai_subtitle(ai_cc_id, base_filename)
            if srt_path:
                self.update_status(f"자막 저장 완료: {os.path.basename(srt_path)}")
            else:
                self.update_status("자막 데이터가 없습니다.")
        except Exception as e:
            logging.error(f"AI subtitle only download error for cid={cid}: {e}", exc_info=True)
            self.update_status(f"자막 다운로드 오류: {e}")
            self.master.after(0, messagebox.showerror, "자막 오류", f"자막 다운로드 중 오류가 발생했습니다:\n{e}")
        finally:
            self.master.after(0, progress_bar.stop)
            self.master.after(0, progress_bar.config, {'mode': 'determinate', 'value': 0})
            self.master.after(0, lambda ct=seminar_data['content_type']: self.on_content_select(event=None, content_type_filter=ct))
        logging.debug("_download_ai_subtitle_only finished.")

    def _start_full_bulk_subtitle_download(self):
        if self.full_bulk_subtitle_thread and self.full_bulk_subtitle_thread.is_alive():
            messagebox.showwarning("진행 중", "이미 일괄 다운로드 중입니다.")
            return
        if not self.seminars_full:
            messagebox.showinfo("목록 없음", "세미나 목록을 먼저 새로고침 해주세요.")
            return
        if not messagebox.askyesno("일괄 자막 다운로드",
                                   f"세미나 {len(self.seminars_full)}개의 자막을 일괄 다운로드합니다.\n계속하시겠습니까?"):
            return
        self.full_bulk_subtitle_thread = threading.Thread(
            target=self._full_bulk_subtitle_worker, daemon=True)
        self.full_bulk_subtitle_thread.start()

    def _full_bulk_subtitle_worker(self):
        items = list(self.seminars_full)
        total = len(items)
        self.master.after(0, self.full_bulk_btn.config, {'state': tk.DISABLED})
        self.master.after(0, self.downloader_progress_bar.config, {'mode': 'determinate', 'maximum': total, 'value': 0})
        try:
            extractor = SeleniumSubtitleExtractor(
                download_dir=DOWNLOAD_DIR,
                user_agent=self.selenium_user_agent.get() if self.selenium_user_agent.get() else None,
                headless=self.selenium_headless_mode.get(),
                driver_path=self.selenium_driver_path.get() if self.selenium_driver_path.get() else None
            )
        except Exception as e:
            self.update_status(f"Selenium 초기화 실패: {e}")
            self.master.after(0, self.full_bulk_btn.config, {'state': tk.NORMAL})
            return
        success_count = skip_count = fail_count = 0
        for idx, item in enumerate(items):
            cid = item.get('id')
            title = self._sanitize_filename(item['title'])
            base_filename = os.path.join(DOWNLOAD_DIR, f"{self._date_prefix(item)}{title}")
            txt_path = f"{base_filename}.txt"
            self.master.after(0, self.downloader_progress_bar.config, {'value': idx})
            self.master.after(0, self.full_bulk_label.config,
                              {'text': f"[{idx+1}/{total}] {item['title'][:30]}..."})
            if os.path.exists(txt_path):
                self.update_status(f"[{idx+1}/{total}] 건너뜀(이미 존재): {title}")
                skip_count += 1
                continue
            self.update_status(f"[{idx+1}/{total}] 자막 다운로드 중: {title}")
            try:
                cont_id = extractor.get_cont_id_from_cid(cid, item.get('sid', ''))
                if not cont_id:
                    fail_count += 1
                    continue
                subtitle_filepath = extractor.download_subtitle(cont_id, title)
                if subtitle_filepath:
                    if subtitle_filepath.lower().endswith(".smi"):
                        plain_text = self.extract_text_from_smi(subtitle_filepath)
                        if plain_text:
                            with open(txt_path, 'w', encoding='utf-8') as f:
                                f.write(plain_text)
                    success_count += 1
                else:
                    fail_count += 1
            except Exception as e:
                logging.error(f"[bulk-full] 자막 오류 cid={cid}: {e}", exc_info=True)
                fail_count += 1
        self.master.after(0, self.downloader_progress_bar.config, {'value': total})
        self.master.after(0, self.full_bulk_label.config, {'text': ""})
        self.master.after(0, self.full_bulk_btn.config, {'state': tk.NORMAL})
        self.update_status(f"일괄 자막 완료 — 성공: {success_count}, 건너뜀: {skip_count}, 실패: {fail_count} / 총 {total}개")
        self.master.after(0, messagebox.showinfo, "완료",
                          f"세미나 자막 일괄 다운로드 완료\n성공: {success_count}  건너뜀: {skip_count}  실패: {fail_count}")

    def _start_press_bulk_subtitle_download(self):
        if self.press_bulk_subtitle_thread and self.press_bulk_subtitle_thread.is_alive():
            messagebox.showwarning("진행 중", "이미 일괄 다운로드 중입니다.")
            return
        if not self.seminars_press:
            messagebox.showinfo("목록 없음", "기자회견 목록을 먼저 새로고침 해주세요.")
            return
        if not messagebox.askyesno("일괄 자막 다운로드",
                                   f"기자회견 {len(self.seminars_press)}개의 자막을 일괄 다운로드합니다.\n계속하시겠습니까?"):
            return
        self.press_bulk_subtitle_thread = threading.Thread(
            target=self._press_bulk_subtitle_worker, daemon=True)
        self.press_bulk_subtitle_thread.start()

    def _press_bulk_subtitle_worker(self):
        items = list(self.seminars_press)
        total = len(items)
        self.master.after(0, self.press_bulk_btn.config, {'state': tk.DISABLED})
        self.master.after(0, self.press_conference_progress_bar.config, {'mode': 'determinate', 'maximum': total, 'value': 0})
        try:
            extractor = SeleniumSubtitleExtractor(
                download_dir=DOWNLOAD_DIR,
                user_agent=self.selenium_user_agent.get() if self.selenium_user_agent.get() else None,
                headless=self.selenium_headless_mode.get(),
                driver_path=self.selenium_driver_path.get() if self.selenium_driver_path.get() else None
            )
        except Exception as e:
            self.update_status(f"Selenium 초기화 실패: {e}")
            self.master.after(0, self.press_bulk_btn.config, {'state': tk.NORMAL})
            return
        success_count = skip_count = fail_count = 0
        for idx, item in enumerate(items):
            cid = item.get('id')
            title = self._sanitize_filename(item['title'])
            base_filename = os.path.join(DOWNLOAD_DIR, f"{self._date_prefix(item)}{title}")
            txt_path = f"{base_filename}.txt"
            self.master.after(0, self.press_conference_progress_bar.config, {'value': idx})
            self.master.after(0, self.press_bulk_label.config,
                              {'text': f"[{idx+1}/{total}] {item['title'][:30]}..."})
            if os.path.exists(txt_path):
                self.update_status(f"[{idx+1}/{total}] 건너뜀(이미 존재): {title}")
                skip_count += 1
                continue
            self.update_status(f"[{idx+1}/{total}] 자막 다운로드 중: {title}")
            try:
                cont_id = extractor.get_cont_id_from_cid(cid, item.get('sid', ''))
                if not cont_id:
                    fail_count += 1
                    continue
                subtitle_filepath = extractor.download_subtitle(cont_id, title)
                if subtitle_filepath:
                    if subtitle_filepath.lower().endswith(".smi"):
                        plain_text = self.extract_text_from_smi(subtitle_filepath)
                        if plain_text:
                            with open(txt_path, 'w', encoding='utf-8') as f:
                                f.write(plain_text)
                    success_count += 1
                else:
                    fail_count += 1
            except Exception as e:
                logging.error(f"[bulk-press] 자막 오류 cid={cid}: {e}", exc_info=True)
                fail_count += 1
        self.master.after(0, self.press_conference_progress_bar.config, {'value': total})
        self.master.after(0, self.press_bulk_label.config, {'text': ""})
        self.master.after(0, self.press_bulk_btn.config, {'state': tk.NORMAL})
        self.update_status(f"일괄 자막 완료 — 성공: {success_count}, 건너뜀: {skip_count}, 실패: {fail_count} / 총 {total}개")
        self.master.after(0, messagebox.showinfo, "완료",
                          f"기자회견 자막 일괄 다운로드 완료\n성공: {success_count}  건너뜀: {skip_count}  실패: {fail_count}")

    def _start_summary_bulk_subtitle_download(self):
        if self.summary_bulk_subtitle_thread and self.summary_bulk_subtitle_thread.is_alive():
            messagebox.showwarning("진행 중", "이미 일괄 다운로드 중입니다.")
            return
        if not self.seminars_summary:
            messagebox.showinfo("목록 없음", "요약영상 목록을 먼저 새로고침 해주세요.")
            return
        if not messagebox.askyesno("일괄 자막 다운로드",
                                   f"요약영상 {len(self.seminars_summary)}개의 자막을 일괄 다운로드합니다.\n계속하시겠습니까?"):
            return
        self.summary_bulk_subtitle_thread = threading.Thread(
            target=self._summary_bulk_subtitle_worker, daemon=True)
        self.summary_bulk_subtitle_thread.start()

    def _summary_bulk_subtitle_worker(self):
        items = list(self.seminars_summary)
        total = len(items)
        self.master.after(0, self.summary_bulk_btn.config, {'state': tk.DISABLED})
        self.master.after(0, self.summary_progress_bar.config, {'mode': 'determinate', 'maximum': total, 'value': 0})

        try:
            extractor = SeleniumSubtitleExtractor(
                download_dir=DOWNLOAD_DIR,
                user_agent=self.selenium_user_agent.get() if self.selenium_user_agent.get() else None,
                headless=self.selenium_headless_mode.get(),
                driver_path=self.selenium_driver_path.get() if self.selenium_driver_path.get() else None
            )
        except Exception as e:
            self.update_status(f"Selenium 초기화 실패: {e}")
            self.master.after(0, self.summary_bulk_btn.config, {'state': tk.NORMAL})
            return

        success_count = 0
        skip_count = 0
        fail_count = 0

        for idx, item in enumerate(items):
            cid = item.get('id')
            title = self._sanitize_filename(item['title'])
            base_filename = os.path.join(DOWNLOAD_DIR, f"{self._date_prefix(item)}{title}")
            txt_path = f"{base_filename}.txt"

            self.master.after(0, self.summary_progress_bar.config, {'value': idx})
            self.master.after(0, self.summary_bulk_label.config,
                              {'text': f"[{idx+1}/{total}] {item['title'][:30]}..."})

            if os.path.exists(txt_path):
                self.update_status(f"[{idx+1}/{total}] 건너뜀(이미 존재): {title}")
                skip_count += 1
                continue

            self.update_status(f"[{idx+1}/{total}] 자막 다운로드 중: {title}")
            try:
                cont_id = extractor.get_cont_id_from_cid(cid, item.get('sid', ''))
                if not cont_id:
                    logging.warning(f"[bulk] ContId 없음: cid={cid}")
                    fail_count += 1
                    continue
                subtitle_filepath = extractor.download_subtitle(cont_id, title)
                if subtitle_filepath:
                    if subtitle_filepath.lower().endswith(".smi"):
                        plain_text = self.extract_text_from_smi(subtitle_filepath)
                        if plain_text:
                            with open(txt_path, 'w', encoding='utf-8') as f:
                                f.write(plain_text)
                    success_count += 1
                else:
                    fail_count += 1
            except Exception as e:
                logging.error(f"[bulk] 자막 오류 cid={cid}: {e}", exc_info=True)
                fail_count += 1

        self.master.after(0, self.summary_progress_bar.config, {'value': total})
        self.master.after(0, self.summary_bulk_label.config, {'text': ""})
        self.master.after(0, self.summary_bulk_btn.config, {'state': tk.NORMAL})
        self.update_status(
            f"일괄 자막 완료 — 성공: {success_count}, 건너뜀: {skip_count}, 실패: {fail_count} / 총 {total}개"
        )
        self.master.after(0, messagebox.showinfo, "완료",
                          f"요약영상 자막 일괄 다운로드 완료\n성공: {success_count}  건너뜀: {skip_count}  실패: {fail_count}")
        logging.debug("_summary_bulk_subtitle_worker finished.")

    def start_download_subtitle_selenium_thread(self, seminar_data):
        logging.debug(f"start_download_subtitle_selenium_thread called for seminar ID: {seminar_data.get('id')}")
        if self.download_thread and self.download_thread.is_alive():
            messagebox.showwarning("진행 중", "이미 다른 파일을 다운로드 중입니다.")
            return

        if not seminar_data:
            messagebox.showerror("오류", "다운로드 정보가 올바르지 않습니다.")
            return

        self.download_thread = threading.Thread(target=self._download_subtitle_with_selenium, args=(seminar_data,), daemon=True)
        self.download_thread.start()
        logging.debug("Download thread started.")

    def _download_subtitle_with_selenium(self, seminar_data):
        logging.debug(f"_download_subtitle_with_selenium started for seminar ID: {seminar_data.get('id')}")
        self.master.after(0, lambda ct=seminar_data['content_type']: self.on_content_select(event=None, content_type_filter=ct))

        cid = seminar_data.get('id')
        title = self._sanitize_filename(seminar_data['title'])
        base_filename = os.path.join(DOWNLOAD_DIR, f"{self._date_prefix(seminar_data)}{title}")

        if not cid:
            self.update_status("오류: 세미나 CID가 없습니다.")
            self.master.after(0, lambda ct=seminar_data['content_type']: self.on_content_select(event=None, content_type_filter=ct))
            return

        self.update_status(f"Selenium으로 자막 검색 중: {title}...")
        _ct = seminar_data.get('content_type', 'FULL')
        if _ct == 'FULL':
            progress_bar = self.downloader_progress_bar
        elif _ct == 'SUMMARY':
            progress_bar = self.summary_progress_bar
        else:
            progress_bar = self.press_conference_progress_bar
        self.master.after(0, progress_bar.config, {'mode': 'indeterminate'})
        self.master.after(0, progress_bar.start)

        extractor = None
        subtitle_filepath = None
        try:
            # Initialize SeleniumSubtitleExtractor with current settings
            extractor = SeleniumSubtitleExtractor(
                download_dir=DOWNLOAD_DIR,
                user_agent=self.selenium_user_agent.get() if self.selenium_user_agent.get() else None,
                headless=self.selenium_headless_mode.get(),
                driver_path=self.selenium_driver_path.get() if self.selenium_driver_path.get() else None
            )
            
            cont_id = extractor.get_cont_id_from_cid(cid, seminar_data.get('sid', ''))
            if cont_id:
                self.update_status(f"ContId ({cont_id}) 추출 완료. 자막 다운로드 시도 중...")
                subtitle_filepath = extractor.download_subtitle(cont_id, title) # Use title for base filename
                if subtitle_filepath:
                    self.update_status(f"자막 다운로드 완료: {os.path.basename(subtitle_filepath)}")
                    logging.debug(f"Subtitle downloaded by Selenium to: {subtitle_filepath}")
                    
                    # --- Add SMI to Text Conversion (if SMI) ---
                    if subtitle_filepath.lower().endswith(".smi"):
                        self.update_status(f"SMI 자막에서 텍스트 추출 중: {os.path.basename(subtitle_filepath)}...")
                        plain_text_content = self.extract_text_from_smi(subtitle_filepath)
                        if plain_text_content:
                            text_filename = str(Path(subtitle_filepath).with_suffix('.txt'))
                            try:
                                with open(text_filename, 'w', encoding='utf-8') as f_txt:
                                    f_txt.write(plain_text_content)
                                self.update_status(f"텍스트 파일 생성 완료: {os.path.basename(text_filename)}")
                                logging.debug(f"Text file created from SMI: {text_filename}")
                            except IOError as e:
                                logging.error(f"Failed to save text file {text_filename}: {e}")
                                self.master.after(0, messagebox.showerror, "텍스트 저장 오류", f"텍스트 파일 저장 중 오류가 발생했습니다:\n{e}")
                        else:
                            self.update_status(f"SMI 자막에서 텍스트 추출 실패: {os.path.basename(subtitle_filepath)}")
                            logging.warning(f"Failed to extract text from SMI after Selenium download: {subtitle_filepath}")
                    # --- End SMI to Text Conversion ---

                else:
                    self.update_status("오류: 자막 다운로드 실패 (Selenium).")
                    logging.error(f"Selenium subtitle download failed for CID: {cid}")
            else:
                self.update_status("오류: ContId 추출 실패 (Selenium).")
                logging.error(f"Selenium failed to extract contId for CID: {cid}")

        except Exception as e:
            self.update_status(f"Selenium 자막 다운로드 중 예외 발생: {e}")
            logging.error(f"Unhandled exception during Selenium subtitle download for CID {cid}: {e}", exc_info=True)
            self.master.after(0, messagebox.showerror, "Selenium 오류", f"Selenium 자막 다운로드 중 오류가 발생했습니다:\n{e}")
        finally:
            self.master.after(0, progress_bar.stop)
            self.master.after(0, progress_bar.config, {'mode': 'determinate', 'value': 0})
            self.master.after(0, lambda ct=seminar_data['content_type']: self.on_content_select(event=None, content_type_filter=ct))
        logging.debug("_download_subtitle_with_selenium finished.")

    # ──────────────────────────────────────────────────────────────
    # 상임위 탭 (w3.assembly.go.kr)
    # ──────────────────────────────────────────────────────────────

    def _on_tab_changed(self, event=None):
        """상임위 탭이 처음 선택될 때 캐시 파일에서 목록을 로드합니다."""
        selected = self.notebook.select()
        if selected == str(self.committee_tab) and not self._committee_loaded:
            self._committee_loaded = True
            self._load_w3_committee_cache()

    def _load_w3_committee_cache(self):
        """저장된 상임위 목록 캐시를 읽어 드롭다운을 채웁니다."""
        if os.path.exists(W3_COMM_JSON):
            try:
                with open(W3_COMM_JSON, 'r', encoding='utf-8') as f:
                    saved = json.load(f)
                self.w3_committee_map = saved
                self._w3_name_to_mc   = {v: k for k, v in saved.items()}
                self.master.after(0, lambda: self._update_w3_comm_dropdown(
                    sorted(saved.values())))
                self.update_status(f"상임위 {len(saved)}개 로드 완료 (캐시).")
                return
            except Exception as e:
                logging.warning(f"상임위 캐시 로드 실패: {e}")

        # 캐시 없으면 서버에서 새로 받기
        self._start_w3_committee_fetch()

    def _load_w3_conf_cache(self):
        """저장된 회의 목록 캐시 파일을 메모리에 로드합니다."""
        if os.path.exists(W3_CONF_JSON):
            try:
                with open(W3_CONF_JSON, 'r', encoding='utf-8') as f:
                    self._w3_conf_json_cache = json.load(f)
            except Exception as e:
                logging.warning(f"회의 목록 캐시 로드 실패: {e}")
                self._w3_conf_json_cache = {}
        else:
            self._w3_conf_json_cache = {}

    def _save_w3_conf_cache(self):
        """메모리의 회의 목록 캐시를 파일에 저장합니다."""
        try:
            with open(W3_CONF_JSON, 'w', encoding='utf-8') as f:
                json.dump(self._w3_conf_json_cache, f, ensure_ascii=False, indent=2)
        except Exception as e:
            logging.warning(f"회의 목록 캐시 저장 실패: {e}")

    def _start_w3_committee_fetch(self):
        """상임위 목록을 서버에서 새로 받아옵니다 (새로고침)."""
        if self.w3_fetch_thread and self.w3_fetch_thread.is_alive():
            return
        self.w3_refresh_btn.config(state=tk.DISABLED)
        self.w3_comm_combo.config(values=[])
        self.w3_comm_var.set('')
        self.update_status("상임위 목록 수집 중 (최대 20초)...")
        self.w3_fetch_thread = threading.Thread(
            target=self._fetch_w3_committees, daemon=True)
        self.w3_fetch_thread.start()

    def _fetch_w3_committees(self):
        """menu 코드 1-120을 병렬 스캔하여 상임위 목록을 수집하고 캐시에 저장합니다."""
        found = {}  # mc -> name

        def check(code):
            try:
                vv = int(time.time())
                url = (f"{W3_BASE}/main/service/list.do"
                       f"?cmd=mainList&menu={code}&vv={vv}")
                r = requests.get(url, headers=W3_HEADERS, timeout=8)
                if r.status_code != 200:
                    return
                data = r.json()
                for c in data.get('commList', []):
                    mc = c.get('mc', '').strip()
                    nm = (c.get('commFullName') or c.get('commName', '')).strip()
                    if mc and nm:
                        found[mc] = nm
                if not data.get('commList') and data.get('confList'):
                    mc = (data.get('mc') or str(code)).strip()
                    nm = (data.get('thisName') or data.get('menuName', '')).strip()
                    if nm and mc and mc not in found:
                        found[mc] = nm
            except Exception:
                pass

        with ThreadPoolExecutor(max_workers=15) as ex:
            list(ex.map(check, range(1, 121)))

        self.w3_committee_map = found
        self._w3_name_to_mc   = {v: k for k, v in found.items()}

        # 캐시 파일 저장
        try:
            with open(W3_COMM_JSON, 'w', encoding='utf-8') as f:
                json.dump(found, f, ensure_ascii=False, indent=2)
        except Exception as e:
            logging.warning(f"상임위 캐시 저장 실패: {e}")

        sorted_names = sorted(found.values())
        self.master.after(0, lambda: self._update_w3_comm_dropdown(sorted_names))
        self.update_status(f"상임위 {len(found)}개 수집 완료.")

    def _update_w3_comm_dropdown(self, names):
        self.w3_comm_combo.config(values=names)
        self.w3_refresh_btn.config(state=tk.NORMAL)

    def _on_w3_committee_selected(self, event=None):
        """드롭다운 선택 시 → 캐시에서 로드."""
        name = self.w3_comm_var.get()
        if not name:
            return
        mc = self._w3_name_to_mc.get(name, '')
        if not mc:
            return

        self.w3_conf_tree.delete(*self.w3_conf_tree.get_children())
        self.w3_clip_listbox.delete(0, tk.END)
        self.w3_movie_data = []
        self._set_w3_dl_state(tk.DISABLED)

        # 회의 목록 캐시가 있으면 즉시 표시
        if not self._w3_conf_json_cache:
            self._load_w3_conf_cache()

        start = self.w3_start_var.get().replace('-', '')
        end   = self.w3_end_var.get().replace('-', '')

        if mc in self._w3_conf_json_cache:
            self.w3_conf_cache = self._w3_conf_json_cache[mc]
            self._apply_w3_date_filter(start, end)
            self.update_status(f"{name} {len(self.w3_conf_cache)}건 (캐시). 새 회의가 있으면 '검색'을 누르세요.")
        else:
            # 캐시 없으면 서버에서 받기
            self._start_w3_conf_fetch(mc, name, start, end)

    def _on_w3_search_clicked(self):
        """검색 버튼 → 항상 서버에서 새로 받음."""
        name = self.w3_comm_var.get()
        if not name:
            return
        mc = self._w3_name_to_mc.get(name, '')
        if not mc:
            return

        self.w3_conf_tree.delete(*self.w3_conf_tree.get_children())
        self.w3_clip_listbox.delete(0, tk.END)
        self.w3_movie_data = []
        self._set_w3_dl_state(tk.DISABLED)

        start = self.w3_start_var.get().replace('-', '')
        end   = self.w3_end_var.get().replace('-', '')
        self._start_w3_conf_fetch(mc, name, start, end)

    def _start_w3_conf_fetch(self, mc, name, start, end):
        self.w3_search_btn.config(state=tk.DISABLED)
        self.update_status(f"회의 목록 수집 중: {name}...")
        t = threading.Thread(target=self._fetch_w3_conferences,
                              args=(mc, name, start, end), daemon=True)
        t.start()

    def _apply_w3_date_filter(self, start, end):
        """self.w3_conf_cache를 날짜 필터 후 트리뷰에 표시합니다."""
        confs = self.w3_conf_cache
        if start or end:
            def in_range(c):
                d = c.get('confDate', '').replace('-', '')
                return (not start or d >= start) and (not end or d <= end)
            confs = [c for c in confs if in_range(c)]

        self.w3_conf_tree.delete(*self.w3_conf_tree.get_children())
        self.w3_conf_data = {}
        for conf in confs:
            date  = conf.get('confDate', '')
            title = conf.get('confTitle', '')
            iid = self.w3_conf_tree.insert('', 'end', values=(date, title))
            self.w3_conf_data[iid] = conf
        self.w3_search_btn.config(state=tk.NORMAL)

    def _fetch_w3_conferences(self, mc, comm_name, start_dt, end_dt):
        """회의 목록 수집.
        22대 국회: movieInfo(ct2+ct3 조합)로 전체 스캔 → 완전한 목록.
        이전 대수: mainList로 최신 21건.
        """
        all_confs = []
        seen = set()
        menu = '30' if int(mc) > 100 else mc

        # ── 22대 국회 전체 스캔 ──────────────────────────────────────
        # 현재 최고 ct2 파악 (mainList에서 샘플로)
        try:
            vv = int(time.time())
            r = requests.get(f"{W3_BASE}/main/service/list.do",
                             params={'cmd': 'mainList', 'menu': menu, 'mc': mc,
                                     'ct1': '22', 'pageSize': '21', 'vv': vv},
                             headers=W3_HEADERS, timeout=15)
            recent22 = r.json().get('confList', [])
            max_ct2 = max((int(c.get('ct2', '0')) for c in recent22), default=440)
        except Exception:
            max_ct2 = 440

        CT1_22_START = 410  # 22대 국회 시작 ct2 (415에 여유 포함)

        def check_meeting_22(combo):
            ct2, ct3 = combo
            try:
                vv = int(time.time())
                r = requests.get(f"{W3_BASE}/main/service/movie.do",
                                 params={'cmd': 'movieInfo', 'mc': mc, 'ct1': '22',
                                         'ct2': str(ct2), 'ct3': f'{ct3:02d}', 'vv': vv},
                                 headers=W3_HEADERS, timeout=10)
                if r.status_code != 200:
                    return None
                data = r.json()
                if not data.get('movieList'):
                    return None
                return {
                    'ct1': '22', 'ct2': str(ct2), 'ct3': f'{ct3:02d}',
                    'mc': mc,
                    'confDate': data.get('confDate', ''),
                    'confOpenTime': data.get('confOpenTime', ''),
                    'confTitle': data.get('confTitle', ''),
                    'sami': data.get('sami', '0'),
                    'minutes': data.get('minutes', '0'),
                }
            except Exception:
                return None

        combos = [(ct2, ct3)
                  for ct2 in range(CT1_22_START, max_ct2 + 3)
                  for ct3 in range(1, 20)]

        self.update_status(f"{comm_name} 22대 전체 스캔 중 ({len(combos)}개 조합)...")

        with ThreadPoolExecutor(max_workers=20) as ex:
            results = list(ex.map(check_meeting_22, combos))

        for conf in results:
            if conf:
                key = f"22_{conf['ct2']}_{conf['ct3']}"
                if key not in seen:
                    seen.add(key)
                    all_confs.append(conf)

        # 22대 mainList 보완: 국정감사 등 ct3가 문자형(AB, AC...)인 회의는 숫자 스캔에서 누락됨
        for c in recent22:
            key = f"{c.get('ct1')}_{c.get('ct2')}_{c.get('ct3')}"
            if key not in seen:
                seen.add(key)
                all_confs.append(c)

        # ── 이전 대수: mainList로 최신 21건 ─────────────────────────
        for ct1 in range(21, 15, -1):
            try:
                vv = int(time.time())
                r = requests.get(f"{W3_BASE}/main/service/list.do",
                                 params={'cmd': 'mainList', 'menu': menu, 'mc': mc,
                                         'ct1': str(ct1), 'pageSize': '21', 'vv': vv},
                                 headers=W3_HEADERS, timeout=15)
                for c in r.json().get('confList', []):
                    key = f"{c.get('ct1')}_{c.get('ct2')}_{c.get('ct3')}"
                    if key not in seen:
                        seen.add(key)
                        all_confs.append(c)
            except Exception:
                continue

        # 날짜 필터 (client-side)
        if start_dt or end_dt:
            def in_range(c):
                d = c.get('confDate', '').replace('-', '')
                return (not start_dt or d >= start_dt) and (not end_dt or d <= end_dt)
            all_confs = [c for c in all_confs if in_range(c)]

        all_confs.sort(key=lambda x: (x.get('confDate', ''), x.get('confOpenTime', '')),
                       reverse=True)

        self.w3_conf_cache = all_confs

        # 회의 목록 캐시 파일에 저장 (날짜 필터 전 전체 목록)
        self._w3_conf_json_cache[mc] = all_confs
        self._save_w3_conf_cache()

        self.master.after(0, lambda: self._apply_w3_date_filter(start_dt, end_dt))
        self.update_status(f"{comm_name} 회의 {len(all_confs)}건 수집 완료.")

    def _populate_w3_conf_tree(self):
        self.w3_conf_tree.delete(*self.w3_conf_tree.get_children())
        self.w3_conf_data = {}
        for c in self.w3_conf_cache:
            date_str = c.get('confDate', 'N/A')
            open_time = c.get('confOpenTime', '')
            display_date = f"{date_str} {open_time}".strip()
            title = c.get('confTitle', 'N/A')
            iid = self.w3_conf_tree.insert("", "end",
                                            values=(display_date, title))
            self.w3_conf_data[iid] = c

    def _on_w3_conf_selected(self, event=None):
        sel = self.w3_conf_tree.selection()
        if not sel:
            return
        conf = self.w3_conf_data.get(sel[0])
        if not conf:
            return
        self.w3_current_conf = conf
        self.w3_clip_listbox.delete(0, tk.END)
        self.w3_movie_data = []
        self._set_w3_dl_state(tk.DISABLED)
        self.update_status(f"클립 목록 수집 중: {conf.get('confTitle', '')}...")
        t = threading.Thread(target=self._fetch_w3_clips, args=(conf,), daemon=True)
        t.start()

    def _fetch_w3_clips(self, conf):
        mc  = conf.get('mc', '')
        ct1 = conf.get('ct1', '')
        ct2 = conf.get('ct2', '')
        ct3 = conf.get('ct3', '')
        try:
            vv  = int(time.time())
            url = (f"{W3_BASE}/main/service/movie.do?cmd=movieInfo"
                   f"&mc={mc}&ct1={ct1}&ct2={ct2}&ct3={ct3}&vv={vv}")
            r = requests.get(url, headers=W3_HEADERS, timeout=15)
            data = r.json()
            movie_list = data.get('movieList', [])
            sami_is    = data.get('sami', '0')
            self.w3_movie_data = movie_list
            self.w3_sami_is    = sami_is
            self.master.after(0, lambda: self._populate_w3_clips(movie_list, sami_is))
        except Exception as e:
            logging.error(f"W3 clip fetch error: {e}", exc_info=True)
            self.update_status(f"클립 목록 수집 오류: {e}")

    def _populate_w3_clips(self, movie_list, sami_is):
        self.w3_clip_listbox.delete(0, tk.END)
        if not movie_list:
            self.w3_clip_listbox.insert(tk.END, "(영상 없음)")
            return
        for m in movie_list:
            play  = m.get('playTime', '')
            title = m.get('movieTitle', '')
            self.w3_clip_listbox.insert(tk.END, f"[{play}]  {title}")
        self._set_w3_dl_state(tk.NORMAL)
        self.update_status(f"클립 {len(movie_list)}개 로드 완료.")

    def _set_w3_dl_state(self, state):
        self.w3_dl_video_btn.config(state=state)
        has_stt = bool(self.stt_password.get())
        sub_state = tk.NORMAL if (state == tk.NORMAL and has_stt) else tk.DISABLED
        self.w3_dl_txt_btn.config(state=sub_state)
        self.w3_dl_all_btn.config(state=sub_state)

    # ── 영상 다운로드 ──

    def _start_w3_video_download(self):
        idxs = list(self.w3_clip_listbox.curselection())
        if not idxs:
            messagebox.showwarning("선택 없음", "다운로드할 클립을 선택해주세요.")
            return
        if self.w3_download_thread and self.w3_download_thread.is_alive():
            messagebox.showwarning("진행 중", "이미 다운로드 중입니다.")
            return
        clips = [self.w3_movie_data[i] for i in idxs if i < len(self.w3_movie_data)]
        self.w3_download_thread = threading.Thread(
            target=self._w3_download_videos,
            args=(self.w3_current_conf, clips, self.w3_quality_var.get()), daemon=True)
        self.w3_download_thread.start()

    def _add_w3_to_cart(self):
        idxs = list(self.w3_clip_listbox.curselection())
        if not idxs:
            messagebox.showwarning("선택 없음", "추가할 클립을 선택해주세요.")
            return
        clips = [self.w3_movie_data[i] for i in idxs if i < len(self.w3_movie_data)]
        quality = self.w3_quality_var.get()
        conf = self.w3_current_conf
        import uuid
        for clip in clips:
            cart_item = {
                'id': str(uuid.uuid4()),
                'type': 'W3',
                'display': f"{conf.get('confDate', '')}  {clip.get('movieTitle', '')}  [{quality}]",
                'conf': conf,
                'clip': clip,
                'quality_name': quality,
            }
            self.download_cart.append(cart_item)
            self.cart_listbox.insert(tk.END, cart_item['display'])
        self.update_status(f"장바구니에 {len(clips)}개 추가됨")

    def _add_content_to_cart(self, content_type_filter):
        import uuid
        if content_type_filter == "FULL":
            tree = self.downloader_tree
            quality_name = self.full_quality_var.get()
        elif content_type_filter == "SUMMARY":
            tree = self.summary_tree
            quality_name = self.summary_quality_var.get()
        else:
            tree = self.press_conference_tree
            quality_name = self.press_quality_var.get()

        selected = tree.selection()
        if not selected:
            return

        quality_key = {'고화질': 'videoFile1', '중화질': 'videoFile2', '저화질': 'videoFile3'}.get(quality_name, 'videoFile1')
        added = 0
        for sel in selected:
            item_id_str = tree.item(sel, 'tags')[0]
            item = self.seminars_by_id.get(item_id_str)
            if not item:
                continue
            cart_item = {
                'id': str(uuid.uuid4()),
                'type': 'VPLATFORM',
                'display': f"{item.get('date', '')[:10]}  {item.get('title', '')}  [{quality_name}]",
                'content_item': item,
                'quality_key': quality_key,
                'quality_name': quality_name,
            }
            self.download_cart.append(cart_item)
            self.cart_listbox.insert(tk.END, cart_item['display'])
            added += 1
        self.update_status(f"장바구니에 {added}개 추가됨")
        self.update_status(f"장바구니 추가: {item.get('title', '')} [{quality_name}]")

    def _remove_selected_from_cart(self):
        for idx in reversed(self.cart_listbox.curselection()):
            self.cart_listbox.delete(idx)
            self.download_cart.pop(idx)

    def _clear_cart(self):
        self.cart_listbox.delete(0, tk.END)
        self.download_cart.clear()

    def _start_cart_download(self):
        if not self.download_cart:
            messagebox.showinfo("장바구니 비어있음", "다운로드할 항목이 없습니다.")
            return
        if self.cart_download_thread and self.cart_download_thread.is_alive():
            messagebox.showwarning("진행 중", "이미 다운로드 중입니다.")
            return
        items = list(self.download_cart)
        mode  = self.cart_mode_var.get()  # 'all' | 'video' | 'subtitle'
        self.cart_download_thread = threading.Thread(
            target=self._cart_download_worker, args=(items, mode), daemon=True)
        self.cart_download_thread.start()

    def _cart_download_worker(self, items, mode='all'):
        total = len(items)
        self.master.after(0, self.cart_download_btn.config, {'state': tk.DISABLED})
        self.master.after(0, self.cart_clear_btn.config, {'state': tk.DISABLED})

        do_video    = mode in ('all', 'video')
        do_subtitle = mode in ('all', 'subtitle')
        has_stt     = bool(self.stt_password.get())

        for i, cart_item in enumerate(items):
            self.master.after(0, self.cart_progress_bar.config,
                              {'value': int(i / total * 100)})
            try:
                if cart_item['type'] == 'VPLATFORM':
                    if do_video:
                        self._download_video(cart_item['quality_key'], cart_item['content_item'])
                        if do_subtitle:
                            pass  # _download_video가 자막도 함께 처리함
                    elif do_subtitle:
                        self._download_video(cart_item['quality_key'], cart_item['content_item'],
                                             subtitle_only=True)
                elif cart_item['type'] == 'W3':
                    if do_video:
                        self._w3_download_videos(
                            cart_item['conf'], [cart_item['clip']], cart_item['quality_name'])
                    if do_subtitle and has_stt:
                        self._w3_download_subtitles(
                            cart_item['conf'], [cart_item['clip']], fmt='txt')
            except Exception as e:
                logging.error(f"장바구니 다운로드 오류: {e}", exc_info=True)
            self.master.after(0, self._remove_cart_item_by_id, cart_item['id'])

        self.master.after(0, self.cart_progress_bar.config, {'value': 100})
        self.master.after(0, self.cart_download_btn.config, {'state': tk.NORMAL})
        self.master.after(0, self.cart_clear_btn.config, {'state': tk.NORMAL})
        self.update_status("장바구니 다운로드 완료")

    def _remove_cart_item_by_id(self, item_id):
        for i, item in enumerate(self.download_cart):
            if item['id'] == item_id:
                self.download_cart.pop(i)
                self.cart_listbox.delete(i)
                break

    def _create_cart_tab(self, parent):
        cart_lf = ttk.LabelFrame(parent, text="다운로드 대기 목록", padding=6)
        cart_lf.pack(fill=tk.BOTH, expand=True, pady=(0, 8))

        vsb = ttk.Scrollbar(cart_lf, orient=tk.VERTICAL)
        self.cart_listbox = tk.Listbox(cart_lf, yscrollcommand=vsb.set,
                                       selectmode=tk.EXTENDED,
                                       font=(self._ui_font[0], 13),
                                       height=15)
        vsb.config(command=self.cart_listbox.yview)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        self.cart_listbox.pack(fill=tk.BOTH, expand=True)

        btn_row = ttk.Frame(parent)
        btn_row.pack(fill=tk.X, pady=(0, 8))

        ttk.Button(btn_row, text="선택 제거",
                   command=self._remove_selected_from_cart).pack(side=tk.LEFT, padx=(0, 4))
        self.cart_clear_btn = ttk.Button(btn_row, text="전체 비우기",
                                         command=self._clear_cart)
        self.cart_clear_btn.pack(side=tk.LEFT, padx=(0, 16))

        ttk.Label(btn_row, text="다운로드:").pack(side=tk.LEFT)
        self.cart_mode_var = tk.StringVar(value='all')
        for label, val in [("영상+자막", "all"), ("영상만", "video"), ("자막만", "subtitle")]:
            ttk.Radiobutton(btn_row, text=label, variable=self.cart_mode_var,
                            value=val).pack(side=tk.LEFT, padx=(4, 0))

        self.cart_download_btn = ttk.Button(btn_row, text="▶ 다운로드 시작",
                                            command=self._start_cart_download)
        self.cart_download_btn.pack(side=tk.RIGHT)

        self.cart_progress_bar = ttk.Progressbar(parent, mode='determinate', length=400)
        self.cart_progress_bar.pack(fill=tk.X, pady=(0, 4))

    # ── 엑셀 업로드 탭 ──────────────────────────────────────────────────────
    def _create_excel_tab(self, parent):
        # 파일 선택 영역
        file_row = ttk.Frame(parent)
        file_row.pack(fill=tk.X, pady=(0, 6))
        ttk.Label(file_row, text="엑셀 파일:").pack(side=tk.LEFT)
        self.excel_path_var = tk.StringVar(value="파일을 선택하세요 (.xlsx)")
        ttk.Entry(file_row, textvariable=self.excel_path_var, state='readonly', width=50).pack(side=tk.LEFT, padx=(6, 4), fill=tk.X, expand=True)
        ttk.Button(file_row, text="파일 선택", command=self._excel_pick_file).pack(side=tk.LEFT)

        # 미리보기 트리뷰
        lf = ttk.LabelFrame(parent, text="엑셀 내용 미리보기", padding=4)
        lf.pack(fill=tk.BOTH, expand=True, pady=(0, 6))
        cols = ('서비스명', '영상 제목', '콘텐츠ID', '재생시간', '등록일시')
        vsb = ttk.Scrollbar(lf, orient=tk.VERTICAL)
        hsb = ttk.Scrollbar(lf, orient=tk.HORIZONTAL)
        self.excel_tree = ttk.Treeview(lf, columns=cols, show='headings',
                                       yscrollcommand=vsb.set, xscrollcommand=hsb.set, height=15)
        vsb.config(command=self.excel_tree.yview)
        hsb.config(command=self.excel_tree.xview)
        widths = {'서비스명': 80, '영상 제목': 360, '콘텐츠ID': 80, '재생시간': 70, '등록일시': 140}
        for c in cols:
            self.excel_tree.heading(c, text=c)
            self.excel_tree.column(c, width=widths.get(c, 100), minwidth=50)
        self.excel_tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        lf.rowconfigure(0, weight=1)
        lf.columnconfigure(0, weight=1)

        # 카운트 레이블
        self.excel_count_label = ttk.Label(parent, text="")
        self.excel_count_label.pack(anchor=tk.W)

        # 다운로드 모드 + 버튼
        ctrl = ttk.Frame(parent)
        ctrl.pack(fill=tk.X, pady=(4, 0))
        self.excel_mode_var = tk.StringVar(value='subtitle')
        for label, val in [("자막만", "subtitle"), ("자막+영상", "all"), ("영상만", "video")]:
            ttk.Radiobutton(ctrl, text=label, variable=self.excel_mode_var, value=val).pack(side=tk.LEFT, padx=(0, 8))
        self.excel_dl_btn = ttk.Button(ctrl, text="다운로드 시작", command=self._start_excel_download, state=tk.DISABLED)
        self.excel_dl_btn.pack(side=tk.LEFT, padx=(12, 0))
        self.excel_dl_label = ttk.Label(ctrl, text="")
        self.excel_dl_label.pack(side=tk.LEFT, padx=(10, 0))

        self.excel_progress_bar = ttk.Progressbar(parent, mode='determinate', length=400)
        self.excel_progress_bar.pack(fill=tk.X, pady=(6, 0))

    def _excel_pick_file(self):
        import platform, subprocess
        logging.debug("_excel_pick_file called")
        path = None
        if platform.system() == 'Darwin':
            # macOS: 네이티브 파일 피커 (tkinter filedialog가 번들 앱에서 안 열리는 문제 우회)
            try:
                script = (
                    'tell application "System Events" to activate\n'
                    'set f to choose file with prompt "엑셀 파일 선택" '
                    'of type {"xlsx", "xls", "com.microsoft.excel.xls"}\n'
                    'return POSIX path of f'
                )
                result = subprocess.run(['osascript', '-e', script],
                                        capture_output=True, text=True, timeout=60)
                logging.debug(f"osascript returncode={result.returncode} stdout={result.stdout!r} stderr={result.stderr!r}")
                if result.returncode == 0:
                    path = result.stdout.strip()
            except Exception as e:
                logging.warning(f"osascript 파일 선택 실패, tkinter 폴백: {e}")

        if not path:
            from tkinter import filedialog
            try:
                root = self.master.winfo_toplevel()
                root.lift()
                root.focus_force()
                path = filedialog.askopenfilename(
                    parent=root,
                    title="엑셀 파일 선택",
                    filetypes=[("Excel 파일", "*.xlsx *.xls"), ("모든 파일", "*.*")]
                )
            except Exception as e:
                messagebox.showerror("오류", f"파일 선택 실패: {e}")
                return

        if not path:
            return
        self.excel_path_var.set(path)
        self._excel_load_file(path)

    def _excel_load_file(self, path):
        logging.debug(f"_excel_load_file called with path={path!r}")
        try:
            import openpyxl
            logging.debug(f"openpyxl version: {openpyxl.__version__}")
        except ImportError:
            messagebox.showerror("오류", "openpyxl 미설치\npip install openpyxl")
            return
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

            # 컬럼 인덱스 자동 탐지
            def col_idx(names):
                for name in names:
                    for i, h in enumerate(headers):
                        if h and name in str(h):
                            return i
                return None

            idx_svc   = col_idx(['서비스명'])
            idx_title = col_idx(['영상 제목', '제목'])
            idx_cid   = col_idx(['콘텐츠ID', 'cid', 'CID'])
            idx_dur   = col_idx(['재생 시간', '재생시간', '시간'])
            idx_date  = col_idx(['등록일시', '날짜', '등록일'])

            if idx_cid is None or idx_title is None:
                messagebox.showerror("오류", f"필수 컬럼(콘텐츠ID, 영상 제목)을 찾을 수 없습니다.\n헤더: {headers}")
                return

            SERVICE_MAP = {'기자회견': 'PRESSCONF', '세미나': 'FULL', '요약영상': 'SUMMARY', '정책세미나': 'FULL'}

            items = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                cid = row[idx_cid] if idx_cid is not None else None
                if not cid:
                    continue
                svc_raw = str(row[idx_svc]).strip() if idx_svc is not None and row[idx_svc] else ''
                content_type = SERVICE_MAP.get(svc_raw, 'PRESSCONF')
                items.append({
                    'cid':          str(int(cid)) if isinstance(cid, float) else str(cid),
                    'title':        str(row[idx_title]).strip() if row[idx_title] else '',
                    'service':      svc_raw,
                    'content_type': content_type,
                    'duration':     str(row[idx_dur]).strip() if idx_dur is not None and row[idx_dur] else '',
                    'date':         str(row[idx_date]).strip() if idx_date is not None and row[idx_date] else '',
                })
            wb.close()
        except Exception as e:
            messagebox.showerror("오류", f"엑셀 읽기 실패: {e}")
            return

        self._excel_items = items

        # 트리뷰 채우기
        for row in self.excel_tree.get_children():
            self.excel_tree.delete(row)
        for it in items:
            self.excel_tree.insert('', 'end', values=(
                it['service'], it['title'], it['cid'], it['duration'], it['date']
            ))

        self.excel_count_label.config(text=f"총 {len(items)}개 항목")
        self.excel_dl_btn.config(state=tk.NORMAL if items else tk.DISABLED)

    def _start_excel_download(self):
        if self.excel_download_thread and self.excel_download_thread.is_alive():
            messagebox.showwarning("진행 중", "이미 다운로드 중입니다.")
            return
        if not self._excel_items:
            messagebox.showinfo("없음", "엑셀 파일을 먼저 불러오세요.")
            return
        if not messagebox.askyesno("다운로드 확인",
                                   f"{len(self._excel_items)}개 항목을 다운로드합니다.\n계속하시겠습니까?"):
            return
        mode = self.excel_mode_var.get()
        self.excel_download_thread = threading.Thread(
            target=self._excel_download_worker, args=(list(self._excel_items), mode), daemon=True)
        self.excel_download_thread.start()

    def _excel_find_video_url(self, item):
        """contentList API로 날짜+제목 검색해 영상 URL 반환. 동일 제목 구분은 등록일시로."""
        date_str = item.get('date', '')
        if not date_str:
            return None
        date_ymd = date_str[:10].replace('-', '')  # 'YYYYMMDD'
        ctype = item['content_type']
        raw_title = item['title']
        payload = json.dumps({
            'contentType': ctype,
            'startDate': date_ymd,
            'endDate': date_ymd,
            'page': 1,
            'limit': 50,
        })
        try:
            r = requests.post(self.content_list_url, headers=self.headers, data=payload, timeout=15)
            r.raise_for_status()
            d = r.json()
            lst = d.get('data', {}).get('list', []) if isinstance(d.get('data'), dict) else (d.get('data') or [])
            for entry in lst:
                if entry.get('title', '').strip() == raw_title.strip():
                    url = entry.get('videoFile3') or entry.get('videoFile2') or entry.get('videoFile1')
                    return url
        except Exception as e:
            logging.warning(f"[excel] 영상 URL 검색 실패: {e}")
        return None

    def _excel_download_worker(self, items, mode):
        total = len(items)
        do_subtitle = mode in ('subtitle', 'all')
        do_video    = mode in ('video', 'all')

        self.master.after(0, self.excel_dl_btn.config, {'state': tk.DISABLED})
        self.master.after(0, self.excel_progress_bar.config,
                          {'mode': 'determinate', 'maximum': total, 'value': 0})

        success = skip = fail = 0

        for idx, item in enumerate(items):
            # 엑셀 콘텐츠ID = aiCC (자막 contId). cid/sid와 다름.
            ai_cc_id = item['cid']
            title    = self._sanitize_filename(item['title'])
            date_prefix = item['date'][:10].replace('-', '') + '_' if item.get('date') else ''
            base  = os.path.join(DOWNLOAD_DIR, f"{date_prefix}{title}")

            self.master.after(0, self.excel_progress_bar.config, {'value': idx})
            self.master.after(0, self.excel_dl_label.config,
                              {'text': f"[{idx+1}/{total}] {item['title'][:28]}..."})
            self.update_status(f"[{idx+1}/{total}] {title}")

            try:
                # ── 자막 다운로드 ── (엑셀 콘텐츠ID를 aiCC contId로 직접 사용)
                if do_subtitle:
                    txt_path = f"{base}.txt"
                    if os.path.exists(txt_path):
                        skip += 1
                    else:
                        ai_done = self._download_ai_subtitle(ai_cc_id, base)
                        if ai_done:
                            success += 1
                        else:
                            fail += 1
                            logging.warning(f"[excel] 자막 없음: aiCC={ai_cc_id} {title}")

                # ── 영상 다운로드 ── (날짜+제목으로 contentList 검색)
                if do_video:
                    ext_suffix = '.mp4'
                    out_path = f"{base}{ext_suffix}"
                    if os.path.exists(out_path):
                        if not do_subtitle:
                            skip += 1
                    else:
                        vid_url = self._excel_find_video_url(item)
                        if vid_url:
                            self._download_with_ffmpeg(vid_url, out_path)
                            if not do_subtitle:
                                success += 1
                        else:
                            if not do_subtitle:
                                fail += 1
                            logging.warning(f"[excel] 영상 URL 없음: {title}")

            except Exception as e:
                logging.error(f"[excel] 오류 aiCC={ai_cc_id}: {e}", exc_info=True)
                fail += 1

        self.master.after(0, self.excel_progress_bar.config, {'value': total})
        self.master.after(0, self.excel_dl_label.config, {'text': ""})
        self.master.after(0, self.excel_dl_btn.config, {'state': tk.NORMAL})
        self.update_status(f"엑셀 다운로드 완료 — 성공: {success}, 건너뜀: {skip}, 실패: {fail} / 총 {total}개")
        self.master.after(0, messagebox.showinfo, "완료",
                          f"엑셀 다운로드 완료\n성공: {success}  건너뜀: {skip}  실패: {fail}")
        if extractor:
            try:
                extractor.close()
            except Exception:
                pass

    def _download_with_ffmpeg(self, url, out_path):
        """ffmpeg로 단일 영상 파일을 다운로드합니다."""
        ffmpeg = self._get_ffmpeg_path()
        if not ffmpeg:
            return
        cmd = [ffmpeg, '-y', '-i', url, '-c', 'copy', out_path]
        try:
            subprocess.run(cmd, check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, timeout=3600)
        except Exception as e:
            logging.error(f"ffmpeg error: {e}")

    def _w3_download_videos(self, conf, clips, quality='고화질'):
        mc  = conf.get('mc', '')
        ct1 = conf.get('ct1', '')
        ct2 = conf.get('ct2', '')
        ct3 = conf.get('ct3', '')
        conf_title = self._sanitize_filename(conf.get('confTitle', 'unknown'))
        conf_date  = conf.get('confDate', '').replace('-', '')

        self.master.after(0, self.w3_progress_bar.config, {'mode': 'indeterminate'})
        self.master.after(0, self.w3_progress_bar.start)

        for clip in clips:
            no         = clip.get('no')
            wv         = clip.get('wv', 1)
            clip_title = self._sanitize_filename(clip.get('movieTitle', f'clip_{no}'))
            filename   = os.path.join(
                DOWNLOAD_DIR, f"{conf_date}_{conf_title}_{clip_title}.mp4")

            if os.path.exists(filename):
                self.update_status(f"이미 존재: {os.path.basename(filename)}")
                continue

            try:
                vv  = int(time.time())
                url = (f"{W3_BASE}/main/service/movie.do?cmd=fileInfo"
                       f"&mc={mc}&ct1={ct1}&ct2={ct2}&ct3={ct3}"
                       f"&no={no}&wv={wv}&xreferer=&vv={vv}")
                r   = requests.get(url, headers=W3_HEADERS, timeout=15)
                fp  = r.json().get('filePath', {})

                # 화질 우선순위 선택
                quality_order = {
                    '고화질': ['720p', '480p', '240p'],
                    '중화질': ['480p', '240p', '720p'],
                    '저화질': ['240p', '480p', '720p'],
                }
                m3u8_path = next(
                    (fp.get(k) for k in quality_order.get(quality, ['720p', '480p', '240p']) if fp.get(k)),
                    None
                )
                if not m3u8_path and isinstance(fp, dict):
                    default_key = fp.get('default', '')
                    m3u8_path   = fp.get(default_key, '')

                if not m3u8_path:
                    self.update_status(f"스트림 URL 없음: {clip_title}")
                    continue

                m3u8_url = ('https:' + m3u8_path
                            if m3u8_path.startswith('//') else m3u8_path)

                self.update_status(f"다운로드 중: {os.path.basename(filename)}...")
                cmd = [self._get_ffmpeg_path(),
                       '-reconnect', '1', '-reconnect_streamed', '1', '-reconnect_delay_max', '30',
                       '-i', m3u8_url,
                       '-c', 'copy', '-bsf:a', 'aac_adtstoasc',
                       '-y', '-loglevel', 'error', filename]
                extra = {}
                if sys.platform == 'win32':
                    extra['creationflags'] = subprocess.CREATE_NO_WINDOW
                subprocess.run(cmd, check=True, capture_output=True,
                               text=True, encoding='utf-8', **extra)

                if os.path.exists(filename):
                    size = os.path.getsize(filename) / (1024 * 1024)
                    self.update_status(
                        f"완료: {os.path.basename(filename)} ({size:.1f}MB)")
                else:
                    self.update_status(f"다운로드 실패: {os.path.basename(filename)}")

            except subprocess.CalledProcessError as e:
                err = (e.stderr.strip().splitlines()[-1]
                       if e.stderr and e.stderr.strip() else str(e))
                self.update_status(f"ffmpeg 오류: {err}")
                self.master.after(0, messagebox.showerror, "다운로드 오류",
                                  f"{clip_title}\n\n{err}")
            except Exception as e:
                logging.error(f"W3 video download error: {e}", exc_info=True)
                self.update_status(f"오류: {e}")

        self.master.after(0, self.w3_progress_bar.stop)
        self.master.after(0, self.w3_progress_bar.config,
                          {'mode': 'determinate', 'value': 0})

    # ── 자막 다운로드 ──

    def _start_w3_subtitle_download(self, fmt='txt'):
        idxs = list(self.w3_clip_listbox.curselection())
        if not idxs:
            messagebox.showwarning("선택 없음", "자막을 받을 클립을 선택해주세요.")
            return
        if self.w3_download_thread and self.w3_download_thread.is_alive():
            messagebox.showwarning("진행 중", "이미 다운로드 중입니다.")
            return
        clips = [self.w3_movie_data[i] for i in idxs if i < len(self.w3_movie_data)]
        self.w3_download_thread = threading.Thread(
            target=self._w3_download_subtitles,
            args=(self.w3_current_conf, clips, fmt), daemon=True)
        self.w3_download_thread.start()

    def _start_w3_full_subtitle_download(self):
        if self.w3_download_thread and self.w3_download_thread.is_alive():
            messagebox.showwarning("진행 중", "이미 다운로드 중입니다.")
            return
        if not self.w3_current_conf:
            messagebox.showwarning("선택 없음", "회의를 선택해주세요.")
            return
        self.w3_download_thread = threading.Thread(
            target=self._w3_download_full_subtitles,
            args=(self.w3_current_conf,), daemon=True)
        self.w3_download_thread.start()

    @staticmethod
    def _match_committee_dirs(comm_name, available_dirs):
        """위원회명 키워드로 서버 디렉토리 목록을 찾습니다. 숫자 suffix 무시."""
        stem_to_dirs = {}
        for d in available_dirs:
            stem = re.sub(r'\d+$', '', d)
            stem_to_dirs.setdefault(stem, []).append(d)

        for dir_stem, keywords in _COMM_DIR_KEYWORDS:
            for kw in keywords:
                if kw and kw in comm_name:
                    if dir_stem in stem_to_dirs:
                        return stem_to_dirs[dir_stem]
        return []

    def _w3_download_full_subtitles(self, conf):
        conf_title = self._sanitize_filename(conf.get('confTitle', 'unknown'))
        conf_date  = conf.get('confDate', '')
        date_str   = conf_date.replace('-', '')

        self.master.after(0, self.w3_progress_bar.config, {'mode': 'indeterminate'})
        self.master.after(0, self.w3_progress_bar.start)

        try:
            import paramiko
        except ImportError:
            self.update_status("오류: paramiko 미설치")
            self.master.after(0, self.w3_progress_bar.stop)
            return

        self.update_status(f"STT 서버 전체 자막 수집 중: {conf_title}...")

        try:
            ssh = paramiko.SSHClient()
            ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            ssh.connect(self.stt_host.get(), port=22,
                        username=self.stt_user.get(),
                        password=self.stt_password.get(), timeout=15)
            sftp = ssh.open_sftp()
        except Exception as e:
            self.update_status(f"SFTP 연결 실패: {e}")
            self.master.after(0, self.w3_progress_bar.stop)
            return

        ssh2  = None
        sftp2 = None

        try:
            d_dash = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:]}"

            def _find_base_full(sftp_conn, base_root):
                for candidate in [date_str, d_dash]:
                    try:
                        p = f"{base_root}/{candidate}"
                        sftp_conn.listdir(p)
                        return p
                    except FileNotFoundError:
                        continue
                return None

            def _collect_segments(sftp_conn, base_path):
                mc        = conf.get('mc', '')
                comm_name = self.w3_committee_map.get(mc, '') or conf.get('confTitle', '')
                available_dirs = sftp_conn.listdir(base_path)
                matched = self._match_committee_dirs(comm_name, available_dirs)
                if matched:
                    subdirs = matched
                    self.update_status(f"위원회 디렉토리: {matched} → 자막 수집 중...")
                else:
                    subdirs = available_dirs
                    logging.warning(f"위원회 디렉토리 매칭 실패({comm_name}), 전체 스캔")

                segs = []
                for subdir in subdirs:
                    dir_path = f"{base_path}/{subdir}"
                    try:
                        all_files = sftp_conn.listdir(dir_path)
                        files = [f for f in all_files if f.endswith('.txt.done')]
                    except Exception:
                        continue
                    for fname in files:
                        fpath = f"{dir_path}/{fname}"
                        try:
                            with sftp_conn.open(fpath, 'rb') as fh:
                                raw = fh.read()
                            for line in raw.decode('utf-8', errors='ignore').split('\n'):
                                line = line.strip()
                                if not line or line == 'HW':
                                    continue
                                try:
                                    item = json.loads(line)
                                except json.JSONDecodeError:
                                    continue
                                if item.get('transcript', '').strip():
                                    segs.append(item)
                        except Exception as e:
                            logging.debug(f"STT file read error {fpath}: {e}")
                return segs

            # 1) 주 서버에서 시도
            all_segments = []
            base_path = _find_base_full(sftp, self.stt_base_path.get())
            if base_path:
                all_segments = _collect_segments(sftp, base_path)
            else:
                logging.warning(f"[STT 전체] 주 서버에서 날짜 디렉토리 없음: {date_str}")

            # 2) 주 서버에서 못 찾으면 보조 서버 (터널링)
            if not all_segments and self.stt_host2.get():
                self.update_status("[STT] 보조 서버 연결 중...")
                try:
                    channel = ssh.get_transport().open_channel(
                        'direct-tcpip',
                        (self.stt_host2.get(), 22),
                        ('127.0.0.1', 0),
                    )
                    ssh2 = paramiko.SSHClient()
                    ssh2.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                    ssh2.connect(
                        self.stt_host2.get(), sock=channel,
                        username=self.stt_user2.get(),
                        password=self.stt_password2.get(),
                        timeout=15,
                    )
                    sftp2 = ssh2.open_sftp()
                    base_root2 = self.stt_base_path2.get() or self.stt_base_path.get()
                    base_path2 = _find_base_full(sftp2, base_root2)
                    if base_path2:
                        all_segments = _collect_segments(sftp2, base_path2)
                    else:
                        logging.warning(f"[STT 전체] 보조 서버에서도 날짜 디렉토리 없음: {date_str}")
                except Exception as e:
                    logging.warning(f"[STT 전체] 보조 서버 연결 실패: {e}")
                    self.update_status(f"보조 서버 연결 실패: {e}")

            if not all_segments:
                self.update_status(f"자막 없음: {conf_title}")
                return

            all_segments.sort(key=lambda x: self._parse_recv_ts(x.get('recv-timestamp', '')))

            content = self._stt_segments_to_txt(all_segments)
            outfile = os.path.join(DOWNLOAD_DIR,
                                   f"{date_str}_{conf_title}_전체.txt")
            with open(outfile, 'w', encoding='utf-8') as f:
                f.write(content)
            self.update_status(
                f"전체 자막 저장: {os.path.basename(outfile)} ({len(all_segments)}줄)")

        except Exception as e:
            logging.error(f"Full STT subtitle error: {e}", exc_info=True)
            self.update_status(f"오류: {e}")
        finally:
            for obj in [sftp2, ssh2, sftp, ssh]:
                try:
                    obj.close()
                except Exception:
                    pass
            self.master.after(0, self.w3_progress_bar.stop)
            self.master.after(0, self.w3_progress_bar.config,
                              {'mode': 'determinate', 'value': 0})

    def _w3_download_subtitles(self, conf, clips, fmt):
        mc  = conf.get('mc', '')
        ct1 = conf.get('ct1', '')
        ct2 = conf.get('ct2', '')
        ct3 = conf.get('ct3', '')
        conf_title = self._sanitize_filename(conf.get('confTitle', 'unknown'))
        conf_date  = conf.get('confDate', '')  # "YYYY-MM-DD" or "YYYYMMDD"

        self.master.after(0, self.w3_progress_bar.config, {'mode': 'indeterminate'})
        self.master.after(0, self.w3_progress_bar.start)

        has_stt = bool(self.stt_password.get())
        date_str = conf_date.replace('-', '')  # YYYYMMDD

        for clip in clips:
            no         = clip.get('no')
            clip_title = self._sanitize_filename(clip.get('movieTitle', f'clip_{no}'))
            base       = os.path.join(
                DOWNLOAD_DIR, f"{date_str}_{conf_title}_{clip_title}")

            written = False

            # ── 1단계: STT 서버에서 자막 시도 ──
            if has_stt:
                real_time = clip.get('realTime', conf.get('confOpenTime', '00:00'))
                play_time = clip.get('playTime', '')
                logging.debug(f"[STT] clip={clip_title} realTime={real_time!r} playTime={play_time!r} keys={list(clip.keys())}")

                if True:  # playTime 유무와 관계없이 시도 (realTime 기본값이면 전체 반환 모드)
                    self.update_status(f"STT 서버 자막 수집 중: {clip_title}...")
                    segments, err = self._fetch_stt_subtitle(conf_date, real_time, play_time or '', conf)
                    if segments:
                        # clip_start_ms 재계산 (SRT 타임스탬프용)
                        rt = real_time.strip()
                        if len(rt) == 5:
                            rt += ':00'
                        dt_kst = datetime.strptime(f"{date_str} {rt}", "%Y%m%d %H:%M:%S")
                        epoch = datetime(1970, 1, 1)
                        clip_start_ms = int((dt_kst - epoch).total_seconds() * 1000) - 9 * 3600 * 1000

                        if fmt == 'srt':
                            content = self._stt_segments_to_srt(segments, clip_start_ms)
                            outfile = base + '_stt.srt'
                        else:
                            content = self._stt_segments_to_txt(segments)
                            outfile = base + '_stt.txt'

                        try:
                            with open(outfile, 'w', encoding='utf-8') as f:
                                f.write(content)
                            self.update_status(
                                f"STT 자막 저장: {os.path.basename(outfile)} ({len(segments)}줄)")
                            written = True
                        except Exception as e:
                            logging.error(f"STT subtitle save error: {e}", exc_info=True)
                    else:
                        self.update_status(f"STT 자막 없음: {clip_title} ({err})")
                        logging.warning(f"STT subtitle not found: {err}, clip={clip_title}")

            if not written:
                self.update_status(f"자막 없음: {clip_title}")

        self.master.after(0, self.w3_progress_bar.stop)
        self.master.after(0, self.w3_progress_bar.config,
                          {'mode': 'determinate', 'value': 0})

    def _date_prefix(self, seminar_data):
        """날짜+시간을 파일명 prefix로 변환 (예: 202603101500_)"""
        raw = seminar_data.get('date', '')
        prefix = raw.replace('-', '').replace(' ', '').replace(':', '')[:12]
        return f"{prefix}_" if prefix else ""

    def _sanitize_filename(self, name):
        name = re.sub(r'[<>:"/\\|?*]', '', name)
        name = name.replace('&#8230;', '').replace('&#34;', '"').replace('&#39;', "'").replace('&#183;', '·') 
        return re.sub(r'\s+', ' ', name)[:150].strip()

    def _check_for_updates(self):
        """GitHub Releases에서 최신 버전을 확인하고 업데이트가 있으면 팝업을 띄웁니다."""
        try:
            resp = requests.get(UPDATE_CHECK_URL, timeout=5,
                                headers={"Accept": "application/vnd.github+json"})
            resp.raise_for_status()
            data = resp.json()
            latest = data.get("tag_name", "").lstrip("v")
            def _ver(v):
                try:
                    return tuple(int(x) for x in v.split('.'))
                except ValueError:
                    return (0,)
            if not latest or _ver(latest) <= _ver(APP_VERSION):
                return

            # 플랫폼별 첨부 파일 탐색
            installer_url = ""
            for asset in data.get("assets", []):
                name = asset["name"]
                if sys.platform == "win32" and name.endswith(".exe"):
                    installer_url = asset["browser_download_url"]
                    break
                if sys.platform == "darwin" and name.endswith(".dmg"):
                    installer_url = asset["browser_download_url"]
                    break

            if not installer_url:
                return

            self.master.after(0, lambda: self._prompt_update(latest, installer_url))
        except Exception as e:
            logging.debug(f"업데이트 확인 실패 (무시): {e}")

    def _prompt_update(self, latest_version, installer_url):
        """업데이트 팝업을 띄우고 사용자가 확인하면 자동 설치합니다."""
        answer = messagebox.askyesno(
            "업데이트 알림",
            f"새 버전이 있습니다.\n\n현재: {APP_VERSION}  →  최신: {latest_version}\n\n"
            f"지금 업데이트하시겠습니까?\n(확인을 누르면 자동으로 설치됩니다)"
        )
        if not answer:
            return
        self.update_status("업데이트 다운로드 중...")
        threading.Thread(target=self._download_and_install_update,
                         args=(installer_url,), daemon=True).start()

    def _download_and_install_update(self, installer_url):
        """인스톨러/DMG를 다운로드하고 플랫폼에 맞게 자동 설치합니다."""
        try:
            resp = requests.get(installer_url, stream=True, timeout=60)
            resp.raise_for_status()

            ext      = ".dmg" if sys.platform == "darwin" else ".exe"
            tmp_name = f"DodioDownloader_Setup{ext}"
            tmp_path = os.path.join(tempfile.gettempdir(), tmp_name)

            total      = int(resp.headers.get('content-length', 0))
            downloaded = 0
            with open(tmp_path, 'wb') as f:
                for chunk in resp.iter_content(chunk_size=8192):
                    if chunk:
                        f.write(chunk)
                        downloaded += len(chunk)
                        if total:
                            pct = int(downloaded / total * 100)
                            self.master.after(0, self.update_status,
                                              f"업데이트 다운로드 중... {pct}%")

            self.master.after(0, self.update_status, "업데이트 설치 중...")

            if sys.platform == "win32":
                # Windows: .bat → 인스톨러 실행 후 재시작
                bat_path = os.path.join(tempfile.gettempdir(), "dodio_update.bat")
                with open(bat_path, 'w', encoding='mbcs') as f:
                    f.write(
                        f'@echo off\r\n'
                        f'timeout /t 3 /nobreak >nul\r\n'
                        f'start /wait "" "{tmp_path}" /VERYSILENT /NORESTART\r\n'
                        f'start "" "{sys.executable}"\r\n'
                    )
                subprocess.Popen(['cmd', '/c', 'start', '', '/b', 'cmd', '/c', bat_path])
                self.master.after(500, self.master.destroy)

            elif sys.platform == "darwin":
                # macOS: DMG 마운트 → .app 교체 → 재시작
                if not getattr(sys, 'frozen', False):
                    # 개발 환경: DMG만 열어줌
                    subprocess.Popen(['open', tmp_path])
                    return

                # sys.executable = .../DodioDownloader.app/Contents/MacOS/DodioDownloader
                app_bundle  = os.path.normpath(
                    os.path.join(os.path.dirname(sys.executable), '..', '..'))
                install_dir = os.path.dirname(app_bundle)
                mnt_path    = os.path.join(tempfile.gettempdir(), "DodioMnt")
                sh_path     = os.path.join(tempfile.gettempdir(), "dodio_update.sh")

                with open(sh_path, 'w') as f:
                    f.write('#!/bin/bash\n')
                    f.write('sleep 3\n')
                    f.write(f'hdiutil attach "{tmp_path}" '
                            f'-mountpoint "{mnt_path}" -nobrowse -quiet\n')
                    f.write(f'cp -R "{mnt_path}/DodioDownloader.app" "{install_dir}/"\n')
                    f.write(f'hdiutil detach "{mnt_path}" -quiet\n')
                    f.write(f'open "{install_dir}/DodioDownloader.app"\n')

                os.chmod(sh_path, 0o755)
                subprocess.Popen(['bash', sh_path])
                self.master.after(500, self.master.destroy)

        except Exception as e:
            logging.error(f"업데이트 다운로드 실패: {e}")
            self.master.after(0, messagebox.showerror, "업데이트 실패",
                              f"다운로드 중 오류가 발생했습니다:\n{e}")

    def _get_ffmpeg_path(self):
        """번들 실행 파일 옆의 ffmpeg(/.exe)를 우선 사용하고, 없으면 PATH에서 찾습니다."""
        ffmpeg_name = 'ffmpeg.exe' if sys.platform == 'win32' else 'ffmpeg'
        candidates = []

        if getattr(sys, 'frozen', False):
            # PyInstaller 빌드
            exe_dir = os.path.dirname(sys.executable)
            candidates.append(os.path.join(exe_dir, ffmpeg_name))
            # PyInstaller 6.x _internal 폴더 대응 (Windows)
            candidates.append(os.path.join(exe_dir, '_internal', ffmpeg_name))
            # macOS .app 번들: Contents/MacOS/ 아래
            candidates.append(os.path.join(exe_dir, '..', 'MacOS', ffmpeg_name))
            # macOS .app 번들: Contents/Frameworks/ 아래
            candidates.append(os.path.join(exe_dir, '..', 'Frameworks', ffmpeg_name))
            # macOS .app 번들: Contents/Resources/ 아래
            candidates.append(os.path.join(exe_dir, '..', 'Resources', ffmpeg_name))
        else:
            # 개발 환경: 스크립트와 같은 폴더
            script_dir = os.path.dirname(os.path.abspath(__file__))
            candidates.append(os.path.join(script_dir, ffmpeg_name))

        for path in candidates:
            if os.path.isfile(path):
                logging.debug(f"ffmpeg found at: {path}")
                return path

        logging.debug(f"ffmpeg not found in: {candidates}, falling back to PATH")
        return 'ffmpeg'

    def _check_ffmpeg(self):
        ffmpeg_path = self._get_ffmpeg_path()
        try:
            result = subprocess.run(
                [ffmpeg_path, '-version'],
                capture_output=True,
                creationflags=subprocess.CREATE_NO_WINDOW if sys.platform == 'win32' else 0
            )
            if result.returncode == 0:
                logging.debug(f"ffmpeg OK: {ffmpeg_path}")
                return True
            logging.error(f"ffmpeg returned {result.returncode}: {result.stderr}")
            return False
        except Exception as e:
            logging.error(f"ffmpeg check failed ({ffmpeg_path}): {e}")
            return False

    def open_download_folder(self):
        if not os.path.exists(DOWNLOAD_DIR):
            os.makedirs(DOWNLOAD_DIR)
        if sys.platform == "win32":
            os.startfile(DOWNLOAD_DIR)
        elif sys.platform == "darwin":
            subprocess.run(["open", DOWNLOAD_DIR])
        else:
            subprocess.run(["xdg-open", DOWNLOAD_DIR])

    def _open_settings_dialog(self):
        """Selenium 설정 다이얼로그를 엽니다."""
        dialog = tk.Toplevel(self.master)
        dialog.title("Selenium 설정")
        dialog.resizable(False, False)
        dialog.transient(self.master)
        dialog.grab_set()

        frame = ttk.Frame(dialog, padding=15)
        frame.pack(fill=tk.BOTH, expand=True)

        self._create_selenium_config_frame(frame)

        ttk.Button(frame, text="닫기", command=dialog.destroy).pack(anchor=tk.E, pady=(10, 0))

        dialog.update_idletasks()
        # 부모 창 중앙에 배치
        pw = self.master.winfo_x() + self.master.winfo_width() // 2
        ph = self.master.winfo_y() + self.master.winfo_height() // 2
        dialog.geometry(f"+{pw - dialog.winfo_width() // 2}+{ph - dialog.winfo_height() // 2}")

    def _create_selenium_config_frame(self, parent):
        """Selenium 설정 UI를 생성합니다."""
        # Headless mode checkbox
        headless_check = ttk.Checkbutton(parent, text="헤드리스 모드 (브라우저 창 숨김)", 
                                         variable=self.selenium_headless_mode)
        headless_check.pack(anchor=tk.W, pady=(0, 5))

        # Chromedriver path entry
        driver_path_frame = ttk.Frame(parent)
        driver_path_frame.pack(fill=tk.X, pady=(0, 5))

        ttk.Label(driver_path_frame, text="WebDriver 경로 (선택 사항, 빈칸시 자동 탐색):").pack(side=tk.LEFT)
        driver_path_entry = ttk.Entry(driver_path_frame, textvariable=self.selenium_driver_path, width=50)
        driver_path_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(5, 0))
        browse_button = ttk.Button(driver_path_frame, text="찾아보기", command=self._browse_webdriver_path)
        browse_button.pack(side=tk.LEFT, padx=(5,0))

        # User-Agent entry
        ua_frame = ttk.Frame(parent)
        ua_frame.pack(fill=tk.X, pady=(0, 5))
        ttk.Label(ua_frame, text="User-Agent (빈칸시 기본값 사용):").pack(side=tk.LEFT)
        ua_entry = ttk.Entry(ua_frame, textvariable=self.selenium_user_agent, width=50)
        ua_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(5,0))

    def _browse_webdriver_path(self):
        filepath = filedialog.askopenfilename(
            title="WebDriver 실행 파일 선택",
            filetypes=(("Executable files", "*.exe"), ("All files", "*.*")),
        )
        if filepath:
            self.selenium_driver_path.set(filepath)

    # ── STT 서버 설정 ──

    def _open_stt_settings_dialog(self):
        """STT 자막 서버 설정 다이얼로그를 엽니다."""
        dialog = tk.Toplevel(self.master)
        dialog.title("STT 자막 서버 설정")
        dialog.resizable(False, False)
        dialog.transient(self.master)
        dialog.grab_set()

        frame = ttk.Frame(dialog, padding=15)
        frame.pack(fill=tk.BOTH, expand=True)

        def row(label, var, show=''):
            f = ttk.Frame(frame)
            f.pack(fill=tk.X, pady=3)
            ttk.Label(f, text=label, width=18, anchor=tk.W).pack(side=tk.LEFT)
            ttk.Entry(f, textvariable=var, width=36, show=show).pack(side=tk.LEFT, fill=tk.X, expand=True)

        ttk.Label(frame, text="▶ 주 서버", font=('', 10, 'bold')).pack(anchor=tk.W, pady=(0, 2))
        row("서버 주소:", self.stt_host)
        row("사용자명:", self.stt_user)
        row("비밀번호:", self.stt_password, show='*')
        row("기본 경로:", self.stt_base_path)

        ttk.Separator(frame, orient='horizontal').pack(fill=tk.X, pady=8)
        ttk.Label(frame, text="▶ 보조 서버 (주 서버 경유 터널)", font=('', 10, 'bold')).pack(anchor=tk.W, pady=(0, 2))
        row("서버 주소:", self.stt_host2)
        row("사용자명:", self.stt_user2)
        row("비밀번호:", self.stt_password2, show='*')
        row("기본 경로:", self.stt_base_path2)

        ttk.Label(frame,
                  text="비밀번호 입력 후 자막 다운로드(SRT/TXT) 버튼이 활성화됩니다.",
                  foreground='gray').pack(anchor=tk.W, pady=(8, 0))

        # 테스트 결과 텍스트
        test_out = tk.Text(frame, height=8, width=55, state=tk.DISABLED,
                           font=('Courier', 9))
        test_out.pack(fill=tk.X, pady=(8, 0))

        def run_test():
            date_input = test_date_var.get().strip().replace('-', '')
            if not date_input:
                messagebox.showwarning("입력 필요", "테스트할 날짜를 입력하세요.", parent=dialog)
                return
            test_btn.config(state=tk.DISABLED, text="테스트 중...")

            def do_test():
                lines = []
                ssh = sftp = ssh2 = sftp2 = None
                try:
                    import paramiko

                    def _test_scan(sftp_conn, base, label):
                        d8 = date_input
                        d_dash = f"{d8[:4]}-{d8[4:6]}-{d8[6:]}"
                        found = None
                        for candidate in [d8, d_dash]:
                            try:
                                p = f"{base}/{candidate}"
                                sftp_conn.listdir(p)
                                found = p
                                break
                            except FileNotFoundError:
                                lines.append(f"[없음] {p}\n")
                        if found is None:
                            return
                        lines.append(f"[OK] {label} 경로: {found}\n")
                        dirs = sftp_conn.listdir(found)
                        lines.append(f"디렉토리: {dirs}\n")
                        for d in dirs:
                            dp = f"{found}/{d}"
                            try:
                                files = sftp_conn.listdir(dp)
                                done_files = [f for f in files if f.endswith('.txt.done')]
                                lines.append(f"\n[{d}] 파일 {len(done_files)}개: {done_files}")
                                for fname in done_files:
                                    fp = f"{dp}/{fname}"
                                    with sftp_conn.open(fp, 'rb') as fh:
                                        head = fh.read(512).decode('utf-8', errors='ignore')
                                    json_lines = [l.strip() for l in head.split('\n') if l.strip() and l.strip() != 'HW']
                                    try:
                                        ts = json.loads(json_lines[0]).get('recv-timestamp', '?') if json_lines else '?'
                                        lines.append(f"  첫 recv-timestamp: {ts}")
                                    except Exception:
                                        lines.append(f"  첫 줄 파싱 실패")
                            except Exception as e:
                                lines.append(f"\n[{d}] 접근 오류: {e}")

                    # 주 서버
                    ssh = paramiko.SSHClient()
                    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                    ssh.connect(self.stt_host.get(), port=22,
                                username=self.stt_user.get(),
                                password=self.stt_password.get(), timeout=10)
                    sftp = ssh.open_sftp()
                    lines.append(f"[OK] 주 서버 연결 성공\n")
                    lines.append(f"\n=== 주 서버 ({self.stt_host.get()}) ===\n")
                    _test_scan(sftp, self.stt_base_path.get(), "주 서버")

                    # 보조 서버 (터널링)
                    if self.stt_host2.get():
                        lines.append(f"\n=== 보조 서버 ({self.stt_host2.get()}) [터널] ===\n")
                        try:
                            channel = ssh.get_transport().open_channel(
                                'direct-tcpip', (self.stt_host2.get(), 22), ('127.0.0.1', 0))
                            ssh2 = paramiko.SSHClient()
                            ssh2.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                            ssh2.connect(self.stt_host2.get(), sock=channel,
                                         username=self.stt_user2.get(),
                                         password=self.stt_password2.get(), timeout=10)
                            sftp2 = ssh2.open_sftp()
                            lines.append(f"[OK] 보조 서버 연결 성공\n")
                            base_root2 = self.stt_base_path2.get() or self.stt_base_path.get()
                            _test_scan(sftp2, base_root2, "보조 서버")
                        except Exception as e:
                            lines.append(f"[오류] 보조 서버 연결 실패: {e}\n")

                except ImportError:
                    lines.append("[오류] paramiko 미설치 (pip install paramiko)")
                except Exception as e:
                    lines.append(f"[오류] {e}")
                finally:
                    for obj in [sftp2, ssh2, sftp, ssh]:
                        try: obj.close()
                        except: pass
                return lines

            def finish(lines):
                result = '\n'.join(lines)
                test_out.config(state=tk.NORMAL)
                test_out.delete('1.0', tk.END)
                test_out.insert(tk.END, result)
                test_out.config(state=tk.DISABLED)
                test_btn.config(state=tk.NORMAL, text="연결 테스트")

            def worker():
                result_lines = do_test()
                dialog.after(0, finish, result_lines)

            threading.Thread(target=worker, daemon=True).start()

        # 날짜 입력 + 테스트 버튼
        test_row = ttk.Frame(frame)
        test_row.pack(fill=tk.X, pady=(6, 0))
        test_date_var = tk.StringVar(value=datetime.today().strftime('%Y%m%d'))
        ttk.Label(test_row, text="테스트 날짜:").pack(side=tk.LEFT)
        ttk.Entry(test_row, textvariable=test_date_var, width=12).pack(side=tk.LEFT, padx=(4, 8))
        test_btn = ttk.Button(test_row, text="연결 테스트", command=run_test)
        test_btn.pack(side=tk.LEFT)

        def on_close():
            self._save_config()
            if self.w3_movie_data:
                self._set_w3_dl_state(tk.NORMAL)
            dialog.destroy()

        ttk.Button(frame, text="확인", command=on_close).pack(anchor=tk.E, pady=(8, 0))

        dialog.update_idletasks()
        pw = self.master.winfo_x() + self.master.winfo_width() // 2
        ph = self.master.winfo_y() + self.master.winfo_height() // 2
        dialog.geometry(f"+{pw - dialog.winfo_width() // 2}+{ph - dialog.winfo_height() // 2}")

    # ── STT 자막 페치 ──

    def _fetch_stt_subtitle(self, conf_date, real_time_str, play_time_str, conf=None):
        """STT 서버 SFTP에서 해당 클립에 해당하는 자막 세그먼트를 반환합니다.

        Returns:
            (segments: list | None, error: str | None)
        """
        try:
            import paramiko
        except ImportError:
            return None, "paramiko 미설치 (pip install paramiko)"

        date_str = conf_date.replace('-', '')  # YYYYMMDD

        # realTime 정규화 → "HH:MM:SS"
        rt = real_time_str.strip()
        if len(rt) == 5:   # "HH:MM"
            rt += ':00'
        elif len(rt) < 8:
            rt = rt.ljust(8, '0')

        real_time_is_default = rt in ('', '00:00:00', '00:0000', '0:00:00')

        # 클립 시작 → Unix ms (KST = UTC+9)
        try:
            dt_kst = datetime.strptime(f"{date_str} {rt}", "%Y%m%d %H:%M:%S")
            epoch = datetime(1970, 1, 1)
            clip_start_ms = int((dt_kst - epoch).total_seconds() * 1000) - 9 * 3600 * 1000
        except ValueError as e:
            if not real_time_is_default:
                return None, f"시각 파싱 오류: {e}"
            clip_start_ms = 0

        # playTime → duration ms
        parts = play_time_str.split(':') if play_time_str else []
        try:
            if len(parts) == 3:
                dur_ms = (int(parts[0]) * 3600 + int(parts[1]) * 60 + int(parts[2])) * 1000
            elif len(parts) == 2:
                dur_ms = (int(parts[0]) * 60 + int(parts[1])) * 1000
            else:
                dur_ms = 0
        except ValueError:
            dur_ms = 0

        # realTime이 기본값이면 dur_ms 체크 없이 전체 반환 모드로 진행
        if dur_ms <= 0 and not real_time_is_default:
            return None, "클립 재생 시간 정보 없음"

        clip_end_ms = clip_start_ms + dur_ms

        # 주 서버 SFTP 연결
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        try:
            ssh.connect(
                self.stt_host.get(), port=22,
                username=self.stt_user.get(),
                password=self.stt_password.get(),
                timeout=15,
            )
            sftp = ssh.open_sftp()
        except Exception as e:
            return None, f"SFTP 연결 실패: {e}"

        ssh2  = None
        sftp2 = None

        try:
            d_dash = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:]}"

            def _find_base(sftp_conn, base_root):
                for candidate in [date_str, d_dash]:
                    try:
                        p = f"{base_root}/{candidate}"
                        sftp_conn.listdir(p)
                        return p
                    except FileNotFoundError:
                        continue
                return None

            def _search_in_base(sftp_conn, base_path):
                try:
                    available_dirs = sftp_conn.listdir(base_path)
                except Exception as e:
                    return None, f"서버 경로 접근 오류: {e}"

                self.update_status(f"[STT] 서버 경로: {base_path}, 디렉토리: {available_dirs}")
                logging.debug(f"[STT] base_path={base_path}, available_dirs={available_dirs}")

                if conf:
                    mc        = conf.get('mc', '')
                    comm_name = self.w3_committee_map.get(mc, '') or conf.get('confTitle', '')
                    matched   = self._match_committee_dirs(comm_name, available_dirs)
                    dirs      = matched if matched else available_dirs
                    self.update_status(f"[STT] 위원회={comm_name!r}, 매칭={matched}, 사용={dirs}")
                    logging.debug(f"[STT] comm_name={comm_name!r}, matched={matched}, dirs={dirs}")
                else:
                    dirs = available_dirs

                best_segments = None

                if not real_time_is_default:
                    for d in dirs:
                        dir_path = f"{base_path}/{d}"
                        try:
                            all_dir_files = sftp_conn.listdir(dir_path)
                            files = [f for f in all_dir_files if f.endswith('.txt.done')]
                            self.update_status(f"[STT] {d}: 전체파일={all_dir_files}, 자막파일={files}")
                            logging.debug(f"[STT] dir={d}, all={all_dir_files}, subtitle_files={files}")
                        except Exception as e:
                            logging.debug(f"[STT] listdir 오류 {dir_path}: {e}")
                            continue

                        for fname in files:
                            fpath = f"{dir_path}/{fname}"
                            try:
                                with sftp_conn.open(fpath, 'rb') as fh:
                                    raw = fh.read()
                                all_lines = [l.strip() for l in raw.decode('utf-8', errors='ignore').split('\n')
                                             if l.strip() and l.strip() != 'HW']
                                if not all_lines:
                                    continue

                                try:
                                    first_ts = self._parse_recv_ts(json.loads(all_lines[0]).get('recv-timestamp', ''))
                                except Exception:
                                    first_ts = 0
                                try:
                                    last_ts = self._parse_recv_ts(json.loads(all_lines[-1]).get('recv-timestamp', ''))
                                except Exception:
                                    last_ts = 0
                                if first_ts and last_ts and (first_ts > clip_end_ms or last_ts < clip_start_ms):
                                    continue

                                segments = []
                                for line in all_lines:
                                    try:
                                        item = json.loads(line)
                                    except json.JSONDecodeError:
                                        continue
                                    recv_ts = self._parse_recv_ts(item.get('recv-timestamp', ''))
                                    if recv_ts and clip_start_ms <= recv_ts < clip_end_ms:
                                        segments.append(item)

                                if segments and (best_segments is None or len(segments) > len(best_segments)):
                                    best_segments = segments
                                    logging.debug(f"STT match: {d}/{fname}, {len(segments)} segments")

                            except Exception as e:
                                logging.debug(f"STT file read error {fpath}: {e}")
                                continue

                if best_segments:
                    best_segments.sort(key=lambda x: self._parse_recv_ts(x.get('recv-timestamp', '')))
                    return best_segments, None

                # 시간 매칭 실패 또는 realTime 기본값 → 디렉토리 전체 반환
                if real_time_is_default:
                    logging.warning("realTime 없음, 디렉토리 전체 반환")
                else:
                    logging.warning("STT 시간 매칭 실패, 디렉토리 전체 반환 fallback")
                all_segs = []
                for d in dirs:
                    dir_path = f"{base_path}/{d}"
                    try:
                        files = [f for f in sftp_conn.listdir(dir_path) if f.endswith('.txt.done')]
                    except Exception:
                        continue
                    for fname in files:
                        fpath = f"{dir_path}/{fname}"
                        try:
                            with sftp_conn.open(fpath, 'rb') as fh:
                                raw = fh.read()
                            for line in raw.decode('utf-8', errors='ignore').split('\n'):
                                line = line.strip()
                                if not line or line == 'HW':
                                    continue
                                try:
                                    item = json.loads(line)
                                except json.JSONDecodeError:
                                    continue
                                if item.get('transcript', '').strip():
                                    all_segs.append(item)
                        except Exception:
                            continue
                if all_segs:
                    all_segs.sort(key=lambda x: self._parse_recv_ts(x.get('recv-timestamp', '')))
                    return all_segs, None
                return None, None  # 자막 없음 (에러 아님)

            # 1) 주 서버에서 시도
            base_path = _find_base(sftp, self.stt_base_path.get())
            if base_path:
                segs, err = _search_in_base(sftp, base_path)
                if segs:
                    return segs, None
                if err:
                    logging.warning(f"[STT] 주 서버 오류: {err}")
            else:
                logging.warning(f"[STT] 주 서버에서 날짜 디렉토리 없음: {date_str}")

            # 2) 보조 서버에서 시도 (주 서버를 점프호스트로 터널링)
            if not self.stt_host2.get():
                return None, "해당 클립에 매칭되는 자막 없음"

            self.update_status("[STT] 보조 서버 연결 중...")
            try:
                channel = ssh.get_transport().open_channel(
                    'direct-tcpip',
                    (self.stt_host2.get(), 22),
                    ('127.0.0.1', 0),
                )
                ssh2 = paramiko.SSHClient()
                ssh2.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                ssh2.connect(
                    self.stt_host2.get(), sock=channel,
                    username=self.stt_user2.get(),
                    password=self.stt_password2.get(),
                    timeout=15,
                )
                sftp2 = ssh2.open_sftp()
            except Exception as e:
                return None, f"보조 서버 연결 실패: {e}"

            base_root2 = self.stt_base_path2.get() or self.stt_base_path.get()
            base_path2 = _find_base(sftp2, base_root2)
            if base_path2:
                segs2, err2 = _search_in_base(sftp2, base_path2)
                if segs2:
                    return segs2, None
                if err2:
                    return None, err2
            else:
                logging.warning(f"[STT] 보조 서버에서도 날짜 디렉토리 없음: {date_str}")

            return None, "해당 클립에 매칭되는 자막 없음"

        finally:
            for obj in [sftp2, ssh2, sftp, ssh]:
                try:
                    obj.close()
                except Exception:
                    pass

    @staticmethod
    def _parse_recv_ts(ts_str):
        """ISO 8601 recv-timestamp 문자열을 UTC ms로 변환합니다."""
        if not ts_str:
            return 0
        try:
            from datetime import timezone as _tz
            dt = datetime.fromisoformat(ts_str)
            return int(dt.timestamp() * 1000)
        except (ValueError, AttributeError):
            return 0

    def _stt_segments_to_srt(self, segments, clip_start_ms):
        """STT 세그먼트 목록을 SRT 문자열로 변환합니다."""
        def to_ts(ms):
            ms = max(0, int(ms))
            h, rem = divmod(ms, 3600000)
            m, rem = divmod(rem, 60000)
            s, ms_part = divmod(rem, 1000)
            return f"{h:02d}:{m:02d}:{s:02d},{ms_part:03d}"

        lines = []
        idx = 1
        for item in segments:
            text = item.get('transcript', '').strip()
            if not text:
                continue
            recv_ts  = self._parse_recv_ts(item.get('recv-timestamp', ''))
            seg_dur  = item.get('segment-duration', 3000)
            start_ms = recv_ts - clip_start_ms
            end_ms   = start_ms + seg_dur
            lines.append(f"{idx}\n{to_ts(start_ms)} --> {to_ts(end_ms)}\n{text}\n")
            idx += 1
        return '\n'.join(lines)

    @staticmethod
    def _stt_segments_to_txt(segments):
        """STT 세그먼트 목록을 순수 텍스트로 변환합니다."""
        return '\n'.join(
            item.get('transcript', '').strip()
            for item in segments
            if item.get('transcript', '').strip()
        )

    @staticmethod
    def extract_text_from_smi(filepath):
        """SMI 파일에서 순수 텍스트를 추출합니다."""
        try:
            import html as html_mod
            with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
            texts = re.findall(r'<SYNC[^>]*><P[^>]*>(.*?)</P>',
                               content, re.IGNORECASE | re.DOTALL)
            lines = []
            for t in texts:
                t = re.sub(r'<[^>]+>', '', t).strip()
                t = html_mod.unescape(t)
                if t and t.lower() != '&nbsp;':
                    lines.append(t)
            return '\n'.join(lines) if lines else None
        except Exception as e:
            logging.error(f"SMI text extraction error: {e}", exc_info=True)
            return None

    def update_status(self, text):
        self.master.after(0, self.status_bar.config, {'text': text})

if __name__ == "__main__":
    try:
        logging.debug("Script execution started.")

        try:
            import ttkbootstrap as ttkbs
            root = ttkbs.Window(themename="flatly")
            root.title("두디오 다운로더")
            root.geometry("1280x860")
            logging.debug("ttkbootstrap Window created (flatly theme).")
        except ImportError:
            root = tk.Tk()
            root.title("두디오 다운로더")
            root.geometry("1280x860")
            logging.debug("ttkbootstrap not found, fallback to tk.Tk.")

        app = SeminarGUI(root)
        logging.debug("SeminarGUI app instance created.")
        root.mainloop()
        logging.debug("Tkinter mainloop exited.")
    except Exception as e:
        logging.error(f"An unhandled exception occurred: {e}", exc_info=True)
        if 'tk' in sys.modules:
            messagebox.showerror("치명적 오류", f"프로그램 실행 중 예기치 않은 오류가 발생했습니다. 자세한 내용은 {LOG_FILE} 파일을 확인해주세요.\n\n오류: {e}")